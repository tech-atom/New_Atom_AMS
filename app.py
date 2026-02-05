from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, send_file
from flask_socketio import SocketIO, emit
import mysql.connector
import os
import random
import time
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
import base64
import cv2
import numpy as np
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
import csv
import secrets
from dotenv import load_dotenv

# Load environment variables
load_dotenv()

# ============================================================================
# FLASK APP INITIALIZATION
# ============================================================================
app = Flask(__name__, static_folder="static", template_folder="templates")

# Generate strong secret key if not in environment
app.secret_key = os.getenv('SECRET_KEY', secrets.token_hex(32))

# Basic Flask configuration
app.config['MAX_CONTENT_LENGTH'] = int(os.getenv('MAX_CONTENT_LENGTH', 52428800))  # 50MB

# Initialize SocketIO with secure configuration
socketio = SocketIO(
    app,
    cors_allowed_origins="*",  # Will be restricted by Cloudflare
    ping_timeout=60,
    ping_interval=25
)

# ============================================================================
# SECURITY INITIALIZATION
# ============================================================================
from security_config import (
    init_security, 
    configure_trusted_proxies,
    sanitize_html,
    validate_input,
    validate_file_upload,
    is_ip_banned,
    record_failed_login,
    clear_failed_login,
    log_security_event,
    require_login,
    require_admin,
    check_ip_ban
)

# Configure trusted proxies (Cloudflare)
configure_trusted_proxies(app)

# Initialize all security features
limiter, csrf = init_security(app)

print("=" * 70)
print("🛡️  ATOM SHAALE AMS - SECURITY ENABLED")
print("=" * 70)

# Configure MIME types for video files
import mimetypes
mimetypes.add_type('video/mp4', '.mp4')
mimetypes.add_type('video/webm', '.webm')
mimetypes.add_type('video/ogg', '.ogv')

# Ensure static files are served with proper headers
@app.after_request
def add_header(response):
    # Allow range requests for video streaming
    if request.path.startswith('/static/'):
        response.headers['Accept-Ranges'] = 'bytes'
    return response

# Database Connection with auto-reconnect and secure configuration
def get_db_connection():
    try:
        # Try DATABASE_URL first (Railway standard), then fall back to individual vars
        database_url = os.getenv('DATABASE_URL')
        if database_url:
            # Parse DATABASE_URL: mysql://user:password@host:port/database
            from urllib.parse import urlparse
            parsed = urlparse(database_url)
            host = parsed.hostname or 'localhost'
            user = parsed.username or 'root'
            password = parsed.password or '12345'
            database = parsed.path.lstrip('/') or 'lms_system'
            port = int(parsed.port or 3306)  # Force integer conversion
        else:
            # Fall back to individual environment variables
            host = os.getenv('DB_HOST', 'localhost')
            user = os.getenv('DB_USER', 'root')
            password = os.getenv('DB_PASSWORD', '12345')
            database = os.getenv('DB_NAME', 'lms_system')
            port = int(os.getenv('DB_PORT') or 3306)
        
        # Debug output
        print(f"[DEBUG] Connecting to MySQL: host={host}, user={user}, database={database}, port={port} (type: {type(port).__name__})")
        
        conn = mysql.connector.connect(
            host=host,
            user=user,
            password=password,
            database=database,
            port=int(port),  # Ensure integer type at connection
            autocommit=False,
            pool_name="mypool",
            pool_size=int(os.getenv('DB_POOL_SIZE', 32)),
            pool_reset_session=True,
            use_pure=True,  # Use pure Python implementation for better security
            ssl_disabled=False  # Enable SSL for database connection in production
        )
        return conn
    except mysql.connector.Error as err:
        log_security_event('database_connection_failed', {'error': str(err)}, 'ERROR')
        print(f"Database connection error: {err}")
        return None

def get_cursor():
    global conn, cursor
    try:
        # Check if connection is alive
        conn.ping(reconnect=True, attempts=3, delay=1)
    except:
        # Reconnect if ping fails
        conn = get_db_connection()
    
    cursor = conn.cursor()
    return cursor

# Perform database migrations at startup
def init_database():
    """Initialize database schema and default data"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Update database schema for correct_option column
        cursor.execute("ALTER TABLE questions MODIFY COLUMN correct_option VARCHAR(10)")
        conn.commit()
        print("✓ Database schema updated: correct_option column extended to VARCHAR(10)")
    except mysql.connector.Error as alter_err:
        if "Duplicate column name" not in str(alter_err):
            print(f"Schema update note: {alter_err}")
    
    try:
        # Create custom registration fields table if not exists
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS registration_fields (
                field_id INT AUTO_INCREMENT PRIMARY KEY,
                field_name VARCHAR(100) NOT NULL,
                field_label VARCHAR(200) NOT NULL,
                field_type ENUM('text', 'email', 'number', 'tel', 'date', 'select', 'textarea') DEFAULT 'text',
                field_options TEXT,
                is_required BOOLEAN DEFAULT TRUE,
                field_order INT DEFAULT 0,
                is_active BOOLEAN DEFAULT TRUE,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        conn.commit()
        print("✓ Registration fields table ready")
        
        # Check if default fields exist (using INSERT IGNORE to avoid duplicates)
        cursor.execute("SELECT COUNT(*) FROM registration_fields WHERE field_name IN ('student_name', 'email', 'password', 'phone', 'roll_number')")
        count = cursor.fetchone()[0]
        
        if count == 0:
            # Insert default registration fields (using INSERT IGNORE to prevent duplicates)
            default_fields = [
                ('student_name', 'Full Name', 'text', None, True, 1),
                ('email', 'Email Address', 'email', None, True, 2),
                ('password', 'Password', 'text', None, True, 3),
                ('phone', 'Phone Number', 'tel', None, False, 4),
                ('roll_number', 'Roll Number', 'text', None, False, 5)
            ]
            cursor.executemany("""
                INSERT IGNORE INTO registration_fields (field_name, field_label, field_type, field_options, is_required, field_order)
                VALUES (%s, %s, %s, %s, %s, %s)
            """, default_fields)
            conn.commit()
            print("✓ Default registration fields created")
            
    except mysql.connector.Error as err:
        print(f"Registration fields table note: {err}")
    
    # Create student_custom_data table for storing custom field values
    try:
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS student_custom_data (
                id INT AUTO_INCREMENT PRIMARY KEY,
                student_id INT NOT NULL,
                field_name VARCHAR(100) NOT NULL,
                field_value TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                FOREIGN KEY (student_id) REFERENCES students(student_id) ON DELETE CASCADE,
                UNIQUE KEY unique_student_field (student_id, field_name),
                INDEX idx_student_id (student_id),
                INDEX idx_field_name (field_name)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        conn.commit()
        print("✓ Student custom data table created")
    except mysql.connector.Error as err:
        print(f"Student custom data table note: {err}")
    
    # Add courses column to exam table for course-based scheduling
    try:
        cursor.execute("""
            ALTER TABLE exam 
            ADD COLUMN courses TEXT COMMENT 'Comma-separated course names'
        """)
        conn.commit()
        print("✓ Exam courses column added")
    except mysql.connector.Error as err:
        if "Duplicate column name" not in str(err):
            print(f"Exam courses column note: {err}")
    
    # Add question_paper_path column for bulk question paper uploads
    try:
        cursor.execute("""
            ALTER TABLE exam 
            ADD COLUMN question_paper_path VARCHAR(500) COMMENT 'Path to uploaded question paper file'
        """)
        conn.commit()
        print("✓ Exam question_paper_path column added")
    except mysql.connector.Error as err:
        if "Duplicate column name" not in str(err):
            print(f"Exam question_paper_path column note: {err}")
    
    # Add exam scheduling columns
    try:
        cursor.execute("""
            ALTER TABLE exam 
            ADD COLUMN start_datetime DATETIME COMMENT 'Exam availability start time'
        """)
        conn.commit()
        print("✓ Exam start_datetime column added")
    except mysql.connector.Error as err:
        if "Duplicate column name" not in str(err):
            print(f"Exam start_datetime column note: {err}")
    
    try:
        cursor.execute("""
            ALTER TABLE exam 
            ADD COLUMN end_datetime DATETIME COMMENT 'Exam deadline/close time'
        """)
        conn.commit()
        print("✓ Exam end_datetime column added")
    except mysql.connector.Error as err:
        if "Duplicate column name" not in str(err):
            print(f"Exam end_datetime column note: {err}")
    
    # Create announcements table
    try:
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS announcements (
                announcement_id INT AUTO_INCREMENT PRIMARY KEY,
                title VARCHAR(200) NOT NULL,
                content TEXT NOT NULL,
                course VARCHAR(100) COMMENT 'Specific course or NULL for all',
                is_pinned BOOLEAN DEFAULT FALSE,
                created_by VARCHAR(100),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                INDEX idx_course (course),
                INDEX idx_pinned (is_pinned)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        conn.commit()
        print("✓ Announcements table created")
    except mysql.connector.Error as err:
        print(f"Announcements table note: {err}")
    
    # Create announcement reads tracking table
    try:
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS announcement_reads (
                read_id INT AUTO_INCREMENT PRIMARY KEY,
                announcement_id INT NOT NULL,
                student_id INT NOT NULL,
                read_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (announcement_id) REFERENCES announcements(announcement_id) ON DELETE CASCADE,
                FOREIGN KEY (student_id) REFERENCES students(student_id) ON DELETE CASCADE,
                UNIQUE KEY unique_read (announcement_id, student_id)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        conn.commit()
        print("✓ Announcement reads table created")
    except mysql.connector.Error as err:
        print(f"Announcement reads table note: {err}")
    
    # Increase selected_option column size for descriptive answers
    try:
        cursor.execute("""
            ALTER TABLE student_responses 
            MODIFY COLUMN selected_option TEXT
        """)
        conn.commit()
        print("✓ Student responses selected_option column expanded to TEXT")
    except mysql.connector.Error as err:
        print(f"Student responses column note: {err}")
    
    # Add submitted_at column if it doesn't exist
    try:
        cursor.execute("""
            ALTER TABLE student_responses 
            ADD COLUMN submitted_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        """)
        conn.commit()
        print("✓ Student responses submitted_at column added")
    except mysql.connector.Error as err:
        if "Duplicate column name" not in str(err):
            print(f"Student responses submitted_at note: {err}")
    
    finally:
        cursor.close()
        conn.close()

# Run database initialization
init_database()

UPLOAD_FOLDER = 'uploads'
MEDIA_FOLDER = 'uploads/media'
STUDENT_RESPONSES_FOLDER = 'uploads/student_responses'
QUESTION_TEMP_FOLDER = 'uploads/temp_questions'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(MEDIA_FOLDER, exist_ok=True)
os.makedirs(STUDENT_RESPONSES_FOLDER, exist_ok=True)
os.makedirs(QUESTION_TEMP_FOLDER, exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MEDIA_FOLDER'] = MEDIA_FOLDER
app.config['STUDENT_RESPONSES_FOLDER'] = STUDENT_RESPONSES_FOLDER
app.config['QUESTION_TEMP_FOLDER'] = QUESTION_TEMP_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024  # 100MB max file size

ALLOWED_EXTENSIONS = {'txt', 'pdf', 'png', 'jpg', 'jpeg', 'gif', 'mp4', 'avi', 'mov', 'mp3', 'wav', 'webm', 'doc', 'docx', 'xlsx', 'xls', 'ppt', 'pptx', 'mkv', 'flv', 'wmv'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# Home Page
@app.route('/')
def home():
    return render_template('index.html')

# Favicon route to prevent 404 errors
@app.route('/favicon.ico')
def favicon():
    return send_file(os.path.join(app.static_folder, 'images', 'atomshaalelogo.png'), mimetype='image/png')

# Test route to verify video file
@app.route('/test-video')
def test_video():
    video_path = os.path.join(app.static_folder, 'images', 'Video_Conversion_to_GIF_Boomerang.mp4')
    exists = os.path.exists(video_path)
    return jsonify({
        'video_exists': exists,
        'video_path': video_path,
        'static_folder': app.static_folder,
        'video_url': url_for('static', filename='images/Video_Conversion_to_GIF_Boomerang.mp4')
    })

# 🔐 Admin Login
@app.route('/login', methods=['GET', 'POST'])
@check_ip_ban  # Check if IP is banned
@limiter.limit("5 per minute")  # Rate limit: 5 attempts per minute
def unified_login():
    # Handle GET requests - redirect to home
    if request.method == 'GET':
        return redirect(url_for('home'))
    
    # Get client IP for logging
    client_ip = request.headers.get('CF-Connecting-IP') or \
                request.headers.get('X-Forwarded-For', '').split(',')[0].strip() or \
                request.remote_addr
    
    identity = request.form.get('identity', '').strip()  # Admin username OR student email
    password = request.form.get('password', '')

    # Input validation
    if not identity or not password:
        log_security_event('login_failed', {'reason': 'missing_credentials', 'ip': client_ip}, 'WARNING')
        flash("Please enter login credentials.", "warning")
        return redirect(url_for('home'))
    
    # Sanitize input (prevent XSS)
    is_valid, identity, error = validate_input(identity, field_type='email' if '@' in identity else 'text', max_length=255)
    if not is_valid:
        log_security_event('login_failed', {'reason': 'invalid_input', 'ip': client_ip}, 'WARNING')
        flash("Invalid input detected.", "danger")
        return redirect(url_for('home'))

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        # ================= ADMIN LOGIN =================
        # Use parameterized query to prevent SQL injection
        cursor.execute("SELECT * FROM admin WHERE username=%s", (identity,))
        admin = cursor.fetchone()

        if admin:
            if check_password_hash(admin['password'], password):
                # Successful admin login
                session.clear()
                session['admin_username'] = admin['username']
                session['role'] = 'admin'
                session['login_time'] = datetime.now().isoformat()
                session['ip_address'] = client_ip
                session.permanent = True  # Make session persistent
                
                clear_failed_login(client_ip)  # Clear failed attempts
                
                log_security_event('login_success', {
                    'user': admin['username'],
                    'role': 'admin',
                    'ip': client_ip
                }, 'INFO')
                
                flash("Welcome back, Admin! Login successful.", "success")
                return redirect(url_for('admin_dashboard'))
            else:
                # Failed admin login
                record_failed_login(client_ip)
                log_security_event('login_failed', {
                    'user': identity,
                    'role': 'admin',
                    'reason': 'invalid_password',
                    'ip': client_ip
                }, 'WARNING')
                
                flash("Invalid admin password.", "danger")
                return redirect(url_for('home'))

        # ================= STUDENT LOGIN =================
        # Use parameterized query to prevent SQL injection
        cursor.execute("SELECT * FROM students WHERE email=%s", (identity,))
        student = cursor.fetchone()

        if not student:
            record_failed_login(client_ip)
            log_security_event('login_failed', {
                'email': identity,
                'reason': 'student_not_found',
                'ip': client_ip
            }, 'WARNING')
            
            flash("Student not registered. Please register first.", "warning")
            return redirect(url_for('home'))

        # Approval checks
        if student['status'] == 'pending':
            log_security_event('login_blocked', {
                'email': identity,
                'reason': 'pending_approval',
                'ip': client_ip
            }, 'INFO')
            
            flash("Your account is pending admin approval.", "info")
            return redirect(url_for('home'))

        if student['status'] == 'rejected':
            log_security_event('login_blocked', {
                'email': identity,
                'reason': 'rejected_account',
                'ip': client_ip
            }, 'WARNING')
            
            flash("Your account has been rejected. Contact admin.", "danger")
            return redirect(url_for('home'))

        # Password verification
        if check_password_hash(student['password'], password):
            # Successful student login
            session.clear()
            session['student_id'] = student['student_id']
            session['student_name'] = student['name']
            session['email'] = student['email']
            session['role'] = 'student'
            session['login_time'] = datetime.now().isoformat()
            session['ip_address'] = client_ip
            session.permanent = True  # Make session persistent
            
            clear_failed_login(client_ip)  # Clear failed attempts
            
            log_security_event('login_success', {
                'user': student['email'],
                'student_id': student['student_id'],
                'role': 'student',
                'ip': client_ip
            }, 'INFO')
            
            flash(f"Welcome, {student['name']}! Login successful.", "success")
            return redirect(url_for('student_dashboard'))

        # Failed student login
        record_failed_login(client_ip)
        log_security_event('login_failed', {
            'email': identity,
            'reason': 'invalid_password',
            'ip': client_ip
        }, 'WARNING')
        
        flash("Invalid email or password.", "danger")
        return redirect(url_for('home'))

    except Exception as e:
        log_security_event('login_error', {
            'error': str(e),
            'ip': client_ip
        }, 'ERROR')
        flash("An error occurred during login. Please try again.", "danger")
        return redirect(url_for('home'))
    
    finally:
        cursor.close()
        conn.close()



# 🎓 Student Registration
@app.route('/register', methods=['GET', 'POST'])
def register():
    conn = get_db_connection()
    cursor = conn.cursor()
    # Fetch active registration fields
    cursor.execute("SELECT * FROM registration_fields WHERE is_active = TRUE ORDER BY field_order, field_id")
    registration_fields = cursor.fetchall()
    
    if request.method == 'POST':
        name = request.form['name']
        email = request.form['email']
        password = generate_password_hash(request.form['password'])  # Hash password
        course = request.form['course']

        try:
            # Insert student record
            cursor.execute("INSERT INTO students (name, email, password, course, status) VALUES (%s, %s, %s, %s, 'pending')", 
                           (name, email, password, course))
            conn.commit()
            
            # Get the newly created student_id
            student_id = cursor.lastrowid
            
            # Save custom field values
            for field in registration_fields:
                field_name = field[1]  # field_name column
                
                # Skip default fields that are already saved in students table
                if field_name in ['student_name', 'email', 'password', 'course']:
                    continue
                
                # Get value from form
                field_value = request.form.get(field_name, '')
                
                # Only save if value is not empty
                if field_value:
                    cursor.execute("""
                        INSERT INTO student_custom_data (student_id, field_name, field_value)
                        VALUES (%s, %s, %s)
                        ON DUPLICATE KEY UPDATE field_value = %s
                    """, (student_id, field_name, field_value, field_value))
            
            conn.commit()
            cursor.close()
            conn.close()
            flash("Registration successful! Wait for admin approval to login.", "info")
            return redirect(url_for('home'))
        except mysql.connector.Error as err:
            cursor.close()
            conn.close()
            flash(f"Error: {err}", "danger")

    cursor.close()
    conn.close()
    return render_template('register.html', fields=registration_fields)

# 🚪 Logout
@app.route('/logout')
def logout():
    session.clear()
    flash("You have been logged out successfully. Come back soon!", "success")
    return redirect(url_for('home'))

# 🏫 Admin Dashboard
@app.route('/admin-dashboard.html')
def admin_dashboard():
    if 'admin_username' not in session:
        flash("Please login as admin first.", "danger")
        return redirect(url_for('home'))
    
    # Fetch dashboard statistics
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Total Students
    cursor.execute("SELECT COUNT(*) FROM students")
    total_students = cursor.fetchone()[0]
    
    # Active Exams
    cursor.execute("SELECT COUNT(*) FROM exam")
    active_exams = cursor.fetchone()[0]
    
   
    # Pending Approvals
    cursor.execute("SELECT COUNT(*) FROM students WHERE status = 'pending'")
    pending_approvals = cursor.fetchone()[0]
    
    cursor.close()
    conn.close()
    
    return render_template('admin-dashboard.html', 
                         total_students=total_students,
                         active_exams=active_exams,
                         pending_approvals=pending_approvals)


# Conduct Exam Page
@app.route('/admin/conduct_exam', methods=['GET'])
def conduct_exam():
    if 'admin_username' not in session:
        flash("Admin access required.", "danger")
        return redirect(url_for('home'))
    return render_template('conduct_exam.html')

# 📢 Announcements Management
@app.route('/admin/announcements')
def admin_announcements():
    if 'admin_username' not in session:
        flash("Admin access required.", "danger")
        return redirect(url_for('home'))
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Fetch all announcements, most recent first
    cursor.execute("""
        SELECT announcement_id, title, content, course, is_pinned, created_at, updated_at
        FROM announcements 
        ORDER BY is_pinned DESC, created_at DESC
    """)
    announcements = cursor.fetchall()
    
    # Get all unique courses from students table
    cursor.execute("SELECT DISTINCT course FROM students WHERE course IS NOT NULL AND course != '' ORDER BY course")
    courses = [row[0] for row in cursor.fetchall()]
    
    cursor.close()
    conn.close()
    
    return render_template('admin_announcements.html', announcements=announcements, courses=courses)

@app.route('/admin/announcements/create', methods=['POST'])
def create_announcement():
    if 'admin_username' not in session:
        flash("Admin access required.", "danger")
        return redirect(url_for('home'))
    
    title = request.form.get('title', '').strip()
    content = request.form.get('content', '').strip()
    course = request.form.get('course', 'All Students')
    is_pinned = 1 if request.form.get('is_pinned') == '1' else 0
    created_by = session['admin_username']
    
    if not title or not content:
        flash("Title and content are required!", "error")
        return redirect(url_for('admin_announcements'))
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        cursor.execute("""
            INSERT INTO announcements (title, content, course, is_pinned, created_by, created_at, updated_at)
            VALUES (%s, %s, %s, %s, %s, NOW(), NOW())
        """, (title, content, course, is_pinned, created_by))
        conn.commit()
        
        # Enhanced flash message with emoji and details
        pin_text = "📌 Pinned" if is_pinned else ""
        course_text = f"for {course}" if course != "All Students" else "to All Students"
        flash(f"🎯 Announcement '{title}' has been created {course_text}! {pin_text}", "success")
    except mysql.connector.Error as err:
        flash(f"Error creating announcement: {err}", "error")
    finally:
        cursor.close()
        conn.close()
    
    return redirect(url_for('admin_announcements'))

@app.route('/admin/announcements/edit/<int:announcement_id>', methods=['POST'])
def edit_announcement(announcement_id):
    if 'admin_username' not in session:
        flash("Admin access required.", "danger")
        return redirect(url_for('home'))
    
    title = request.form.get('title', '').strip()
    content = request.form.get('content', '').strip()
    course = request.form.get('course', 'All Students')
    is_pinned = 1 if request.form.get('is_pinned') == '1' else 0
    
    if not title or not content:
        flash("Title and content are required!", "error")
        return redirect(url_for('admin_announcements'))
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        cursor.execute("""
            UPDATE announcements 
            SET title = %s, content = %s, course = %s, is_pinned = %s, updated_at = NOW()
            WHERE announcement_id = %s
        """, (title, content, course, is_pinned, announcement_id))
        conn.commit()
        flash("Announcement updated successfully!", "success")
    except mysql.connector.Error as err:
        flash(f"Error updating announcement: {err}", "error")
    finally:
        cursor.close()
        conn.close()
    
    return redirect(url_for('admin_announcements'))

@app.route('/admin/announcements/delete/<int:announcement_id>', methods=['POST'])
def delete_announcement(announcement_id):
    if 'admin_username' not in session:
        flash("Admin access required.", "danger")
        return redirect(url_for('home'))
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Delete announcement (reads will be deleted automatically due to foreign key cascade)
        cursor.execute("DELETE FROM announcements WHERE announcement_id = %s", (announcement_id,))
        conn.commit()
        flash("Announcement deleted successfully!", "success")
    except mysql.connector.Error as err:
        flash(f"Error deleting announcement: {err}", "error")
    finally:
        cursor.close()
        conn.close()
    
    return redirect(url_for('admin_announcements'))

@app.route('/admin/announcements/toggle-pin/<int:announcement_id>', methods=['POST'])
def toggle_pin_announcement(announcement_id):
    if 'admin_username' not in session:
        return jsonify({'success': False, 'message': 'Admin access required'}), 403
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Toggle pin status
        cursor.execute("UPDATE announcements SET is_pinned = NOT is_pinned, updated_at = NOW() WHERE announcement_id = %s", (announcement_id,))
        conn.commit()
        
        # Get new pin status
        cursor.execute("SELECT is_pinned FROM announcements WHERE announcement_id = %s", (announcement_id,))
        result = cursor.fetchone()
        is_pinned = result[0] if result else 0
        
        return jsonify({'success': True, 'is_pinned': is_pinned})
    except mysql.connector.Error as err:
        return jsonify({'success': False, 'message': str(err)}), 500
    finally:
        cursor.close()
        conn.close()

# 🎓 Student Dashboard
@app.route('/student-dashboard.html')
def student_dashboard():
    if 'student_id' not in session:
        flash("Please log in first.", "warning")
        return redirect(url_for('unified_login'))

    student_id = session['student_id']

    conn = get_db_connection()
    cursor = conn.cursor()
    

    # Get student's course
    cursor.execute("SELECT course FROM students WHERE student_id = %s", (student_id,))
    student_course_result = cursor.fetchone()
    student_course = student_course_result[0] if student_course_result and student_course_result[0] else None

    # Fetch exams for student's course (or exams assigned to 'All Courses') with scheduling info
    if student_course:
        cursor.execute("""
            SELECT exam_id, exam_title, subject_name, courses, start_datetime, end_datetime
            FROM exam 
            WHERE courses IS NULL 
               OR courses = '' 
               OR courses LIKE '%All Courses%'
               OR FIND_IN_SET(%s, REPLACE(courses, ', ', ',')) > 0
            ORDER BY created_at DESC
        """, (student_course,))
    else:
        # If student has no course, show only "All Courses" exams
        cursor.execute("""
            SELECT exam_id, exam_title, subject_name, courses, start_datetime, end_datetime
            FROM exam 
            WHERE courses IS NULL 
               OR courses = '' 
               OR courses LIKE '%All Courses%'
            ORDER BY created_at DESC
        """)
    raw_exams = cursor.fetchall()  # Fetch the latest exams
    
    # Check which exams the student has already taken
    cursor.execute("SELECT DISTINCT exam_id FROM student_performance WHERE student_id = %s", (student_id,))
    taken_exams = [row[0] for row in cursor.fetchall()]

    cursor.close()
    conn.close()
    
    # Process exams with scheduling status
    from datetime import datetime
    current_time = datetime.now()
    exams = []
    
    for exam in raw_exams:
        exam_id, exam_title, subject_name, courses, start_datetime, end_datetime = exam
        exam_data = {
            'exam_id': exam_id,
            'exam_title': exam_title,
            'subject_name': subject_name,
            'courses': courses,
            'start_datetime': start_datetime,
            'end_datetime': end_datetime,
            'status': 'available',  # available, upcoming, expired
            'time_remaining': None,
            'countdown_seconds': None
        }
        
        # Determine exam status based on schedule
        if start_datetime and end_datetime:
            if current_time < start_datetime:
                exam_data['status'] = 'upcoming'
                time_diff = start_datetime - current_time
                exam_data['countdown_seconds'] = int(time_diff.total_seconds())
                days = time_diff.days
                hours = time_diff.seconds // 3600
                minutes = (time_diff.seconds % 3600) // 60
                if days > 0:
                    exam_data['time_remaining'] = f"{days}d {hours}h"
                elif hours > 0:
                    exam_data['time_remaining'] = f"{hours}h {minutes}m"
                else:
                    exam_data['time_remaining'] = f"{minutes}m"
            elif current_time > end_datetime:
                exam_data['status'] = 'expired'
            else:
                exam_data['status'] = 'available'
        
        exams.append(exam_data)
    
    # Filter out completed exams for display - only show available/upcoming
    available_exams = [exam for exam in exams if exam['exam_id'] not in taken_exams and exam['status'] != 'expired']
    available_exams_count = len([e for e in available_exams if e['status'] == 'available'])
    # Pending means all not-yet-attempted exams (both available and upcoming)
    pending_exams_count = len(available_exams)
    
    # Get student name from session
    student_name = session.get('student_name', 'Student')
    
    # Fetch unread announcements for popup
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Get student's course for announcements
    cursor.execute("SELECT course FROM students WHERE student_id = %s", (student_id,))
    student_course_result = cursor.fetchone()
    student_course = student_course_result[0] if student_course_result and student_course_result[0] else None
    
    # Fetch unread announcements
    if student_course:
        cursor.execute("""
            SELECT a.announcement_id, a.title, a.content, a.course, a.is_pinned, a.created_at
            FROM announcements a
            LEFT JOIN announcement_reads ar ON a.announcement_id = ar.announcement_id AND ar.student_id = %s
            WHERE (a.course = 'All Students' OR a.course = %s)
              AND ar.read_at IS NULL
            ORDER BY a.is_pinned DESC, a.created_at DESC
        """, (student_id, student_course))
    else:
        cursor.execute("""
            SELECT a.announcement_id, a.title, a.content, a.course, a.is_pinned, a.created_at
            FROM announcements a
            LEFT JOIN announcement_reads ar ON a.announcement_id = ar.announcement_id AND ar.student_id = %s
            WHERE a.course = 'All Students'
              AND ar.read_at IS NULL
            ORDER BY a.is_pinned DESC, a.created_at DESC
        """, (student_id,))
    
    unread_announcements = cursor.fetchall()
    cursor.close()
    conn.close()
    
    return render_template('student-dashboard-new.html', 
                         available_exams=available_exams,
                         available_exams_count=available_exams_count,
                         pending_exams_count=pending_exams_count,
                         taken_exams=taken_exams,
                         student_name=student_name,
                         unread_announcements=unread_announcements)

# 🗑️ Admin - Delete Exam
@app.route('/admin/delete_exam/<int:exam_id>', methods=['POST'])
def delete_exam(exam_id):
    if 'admin_username' not in session:
        flash("Admin access required.", "danger")
        return redirect(url_for('home'))
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        # Delete exam (questions will be deleted automatically due to foreign key cascade)
        cursor.execute("DELETE FROM exam WHERE exam_id = %s", (exam_id,))
        conn.commit()
        flash("Exam deleted successfully!", "success")
    except mysql.connector.Error as err:
        flash(f"Error deleting exam: {err}", "danger")
    finally:
        cursor.close()
        conn.close()
    return redirect(url_for('manage_exams'))

# � Student Announcements
@app.route('/student/announcements')
def student_announcements():
    if 'student_id' not in session:
        flash("Please log in first.", "warning")
        return redirect(url_for('login'))
    
    student_id = session['student_id']
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Get student's course
    cursor.execute("SELECT course FROM students WHERE student_id = %s", (student_id,))
    student_course_result = cursor.fetchone()
    student_course = student_course_result[0] if student_course_result and student_course_result[0] else None
    
    # Fetch announcements for student's course or "All Students"
    if student_course:
        cursor.execute("""
            SELECT a.announcement_id, a.title, a.content, a.course, a.is_pinned, a.created_at, a.updated_at,
                   ar.read_at IS NOT NULL as is_read
            FROM announcements a
            LEFT JOIN announcement_reads ar ON a.announcement_id = ar.announcement_id AND ar.student_id = %s
            WHERE a.course = 'All Students' OR a.course = %s
            ORDER BY a.is_pinned DESC, a.created_at DESC
        """, (student_id, student_course))
    else:
        cursor.execute("""
            SELECT a.announcement_id, a.title, a.content, a.course, a.is_pinned, a.created_at, a.updated_at,
                   ar.read_at IS NOT NULL as is_read
            FROM announcements a
            LEFT JOIN announcement_reads ar ON a.announcement_id = ar.announcement_id AND ar.student_id = %s
            WHERE a.course = 'All Students'
            ORDER BY a.is_pinned DESC, a.created_at DESC
        """, (student_id,))
    
    announcements = cursor.fetchall()
    
    # Count unread announcements
    unread_count = sum(1 for a in announcements if not a[7])
    
    cursor.close()
    conn.close()
    
    return render_template('student_announcements.html', 
                         announcements=announcements, 
                         unread_count=unread_count,
                         student_name=session.get('student_name', 'Student'))

@app.route('/student/announcements/mark-read/<int:announcement_id>', methods=['POST'])
def mark_announcement_read(announcement_id):
    if 'student_id' not in session:
        return jsonify({'success': False, 'message': 'Not logged in'}), 401
    
    student_id = session['student_id']
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Check if already marked as read
        cursor.execute("""
            SELECT read_id FROM announcement_reads 
            WHERE announcement_id = %s AND student_id = %s
        """, (announcement_id, student_id))
        
        if not cursor.fetchone():
            # Mark as read
            cursor.execute("""
                INSERT INTO announcement_reads (announcement_id, student_id, read_at)
                VALUES (%s, %s, NOW())
            """, (announcement_id, student_id))
            conn.commit()
        
        return jsonify({'success': True})
    except mysql.connector.Error as err:
        return jsonify({'success': False, 'message': str(err)}), 500
    finally:
        cursor.close()
        conn.close()

@app.route('/student/announcements/unread-count')
def get_unread_count():
    if 'student_id' not in session:
        return jsonify({'count': 0})
    
    student_id = session['student_id']
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Get student's course
    cursor.execute("SELECT course FROM students WHERE student_id = %s", (student_id,))
    student_course_result = cursor.fetchone()
    student_course = student_course_result[0] if student_course_result and student_course_result[0] else None
    
    # Count unread announcements
    if student_course:
        cursor.execute("""
            SELECT COUNT(*) FROM announcements a
            LEFT JOIN announcement_reads ar ON a.announcement_id = ar.announcement_id AND ar.student_id = %s
            WHERE (a.course = 'All Students' OR a.course = %s) AND ar.read_at IS NULL
        """, (student_id, student_course))
    else:
        cursor.execute("""
            SELECT COUNT(*) FROM announcements a
            LEFT JOIN announcement_reads ar ON a.announcement_id = ar.announcement_id AND ar.student_id = %s
            WHERE a.course = 'All Students' AND ar.read_at IS NULL
        """, (student_id,))
    
    count = cursor.fetchone()[0]
    
    cursor.close()
    conn.close()
    
    return jsonify({'count': count})

# �📋 Admin - View All Exams
# 📊 Student Progress Tracking
@app.route('/student/progress')
def student_progress():
    if 'student_id' not in session:
        flash("Please log in first.", "warning")
        return redirect(url_for('unified_login'))
    
    student_id = session['student_id']
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Get student info
    cursor.execute("SELECT name, course FROM students WHERE student_id = %s", (student_id,))
    student_info = cursor.fetchone()
    student_name = student_info[0] if student_info else "Student"
    student_course = student_info[1] if student_info else None
    
    # Get all exam results with dates (only show results where scores are not restricted)
    cursor.execute("""
        SELECT 
            e.exam_title,
            e.subject_name,
            sp.score,
            sp.total_questions,
            sp.recorded_at,
            sp.exam_id
        FROM student_performance sp
        JOIN exam e ON sp.exam_id = e.exam_id
        WHERE sp.student_id = %s
        AND COALESCE(e.show_scores, 1) = 1
        ORDER BY sp.recorded_at ASC
    """, (student_id,))
    
    results = cursor.fetchall()
    
    # Calculate statistics
    total_exams = len(results)
    
    if total_exams > 0:
        # Calculate percentages and prepare data for charts
        exam_data = []
        subject_performance = {}
        
        for result in results:
            exam_title, subject, score, total, recorded_at, exam_id = result
            percentage = round(float(score), 2)
            
            exam_data.append({
                'title': exam_title,
                'subject': subject,
                'score': score,
                'total': total,
                'percentage': percentage,
                'date': recorded_at.strftime('%b %d, %Y'),
                'exam_id': exam_id
            })
            
            # Track subject-wise performance
            if subject not in subject_performance:
                subject_performance[subject] = {'total_score': 0, 'count': 0}
            subject_performance[subject]['total_score'] += percentage
            subject_performance[subject]['count'] += 1
        
        # Calculate average score and improvement trend
        average_score = round(sum([e['percentage'] for e in exam_data]) / total_exams, 2)
        
        # Get weak and strong subjects
        subject_averages = []
        for subject, data in subject_performance.items():
            avg_percentage = round((data['total_score'] / data['count']), 2) if data['count'] > 0 else 0
            subject_averages.append({
                'subject': subject,
                'percentage': avg_percentage,
                'exams_taken': data['count']
            })
        
        subject_averages.sort(key=lambda x: x['percentage'])
        weak_subjects = subject_averages[:3]  # Bottom 3
        strong_subjects = sorted(subject_averages, key=lambda x: x['percentage'], reverse=True)[:3]  # Top 3
        
        # Generate study recommendations
        recommendations = []
        if average_score < 50:
            recommendations.append("Focus on understanding fundamental concepts before attempting advanced topics.")
            recommendations.append("Consider forming study groups or seeking help from instructors.")
        elif average_score < 70:
            recommendations.append("Review incorrect answers to identify patterns in mistakes.")
            recommendations.append("Practice more questions in your weak areas.")
        else:
            recommendations.append("Excellent performance! Focus on maintaining consistency.")
            recommendations.append("Challenge yourself with advanced practice materials.")
        
        # Add subject-specific recommendations
        if weak_subjects:
            recommendations.append(f"Prioritize studying {weak_subjects[0]['subject']} - allocate extra time for practice.")
        
    else:
        exam_data = []
        average_score = 0
        weak_subjects = []
        strong_subjects = []
        recommendations = ["Take your first exam to start tracking your progress!"]
    
    cursor.close()
    conn.close()
    
    # Prepare subject-wise data for new template
    subject_wise = []
    for subject, data in subject_performance.items() if total_exams > 0 else []:
        avg_score = round((data['total_score'] / data['count']), 2) if data['count'] > 0 else 0
        subject_wise.append({
            'subject': subject,
            'avg_score': avg_score,
            'exam_count': data['count']
        })
    subject_wise.sort(key=lambda x: x['avg_score'], reverse=True)
    
    # Get top score
    top_score = max([e['percentage'] for e in exam_data]) if exam_data else 0
    
    return render_template('student_progress_ultra.html',
                         student_name=student_name,
                         course=student_course,
                         exam_data=exam_data,
                         total_exams=total_exams,
                         average_score=average_score,
                         subject_wise=subject_wise,
                         top_score=top_score,
                         recommendations=recommendations)

@app.route('/admin/manage_exams')
def manage_exams():
    if 'admin_username' not in session:
        flash("Admin access required.", "danger")
        return redirect(url_for('home'))
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT e.exam_id, e.exam_title, e.subject_name, e.courses, COUNT(q.question_id) as question_count, e.created_at, 
               COALESCE(e.show_scores, 1) as show_scores
        FROM exam e
        LEFT JOIN questions q ON e.exam_id = q.exam_id
        GROUP BY e.exam_id, e.exam_title, e.subject_name, e.courses, e.created_at, e.show_scores
        ORDER BY e.created_at DESC
    """)
    exams = cursor.fetchall()
    
    # Debug: print exam details
    print("DEBUG MANAGE EXAMS - Exams with question counts:")
    for exam in exams:
        print(f"  Exam ID: {exam[0]}, Title: {exam[1]}, Subject: {exam[2]}, Questions: {exam[3]}")
    
    cursor.close()
    conn.close()
    return render_template('manage_exams.html', exams=exams)

# Edit Exam
@app.route('/admin/edit_exam/<int:exam_id>', methods=['GET', 'POST'])
def edit_exam(exam_id):
    if 'admin_username' not in session:
        flash("Admin access required.", "danger")
        return redirect(url_for('home'))
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    if request.method == 'POST':
        exam_title = request.form.get("exam_title")
        subject = request.form.get("subject")
        time_limit = request.form.get("time_limit")  # Get time limit
        
        questions = request.form.getlist("question[]")
        question_types = request.form.getlist("question_type[]")
        options_a = request.form.getlist("option_a[]")
        options_b = request.form.getlist("option_b[]")
        options_c = request.form.getlist("option_c[]")
        options_d = request.form.getlist("option_d[]")
        correct_options = request.form.getlist("correct_option[]")
        tf_correct = request.form.getlist("tf_correct[]")
        explanations = request.form.getlist("explanation[]")
        question_images = request.files.getlist("question_image[]")
        
        print(f"DEBUG EDIT EXAM - Exam ID: {exam_id}")
        print(f"DEBUG EDIT EXAM - Questions count: {len(questions)}") 
        
        try:
            # Convert time_limit to int or None
            time_limit_int = int(time_limit) if time_limit and time_limit.strip() else None
            
            # Update exam details
            cursor.execute("UPDATE exam SET exam_title = %s, subject_name = %s, time_limit = %s WHERE exam_id = %s",
                         (exam_title, subject, time_limit_int, exam_id))
            
            # Delete all existing questions for this exam
            cursor.execute("DELETE FROM questions WHERE exam_id = %s", (exam_id,))
            
            # Insert updated questions
            insert_question_query = """
            INSERT INTO questions (exam_id, question_text, question_type, option_a, option_b, option_c, option_d, correct_option, explanation, media_path) 
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """
            
            for i in range(len(questions)):
                q_type = question_types[i] if i < len(question_types) else 'mcq'
                image_path = None
                
                # Handle image upload
                if q_type == 'image_mcq' and i < len(question_images) and question_images[i].filename:
                    image_file = question_images[i]
                    filename = secure_filename(image_file.filename)
                    unique_filename = f"exam_{exam_id}_q_{i+1}_{filename}"
                    image_path = os.path.join('static', 'question_images', unique_filename)
                    os.makedirs(os.path.join('static', 'question_images'), exist_ok=True)
                    image_file.save(image_path)
                
                # Handle different question types
                if q_type == 'video_response' or q_type == 'descriptive':
                    cursor.execute(insert_question_query, (
                        exam_id, questions[i], q_type, None, None, None, None, None, 
                        explanations[i] if i < len(explanations) else None, image_path
                    ))
                elif q_type == 'true_false':
                    correct_ans = tf_correct[i] if i < len(tf_correct) else None
                    cursor.execute(insert_question_query, (
                        exam_id, questions[i], q_type, 'True', 'False', None, None, 
                        correct_ans, explanations[i] if i < len(explanations) else None, image_path
                    ))
                else:
                    cursor.execute(insert_question_query, (
                        exam_id, questions[i], q_type, 
                        options_a[i] if i < len(options_a) else None,
                        options_b[i] if i < len(options_b) else None,
                        options_c[i] if i < len(options_c) else None,
                        options_d[i] if i < len(options_d) else None,
                        correct_options[i] if i < len(correct_options) else None,
                        explanations[i] if i < len(explanations) else None,
                        image_path
                    ))
            
            conn.commit()
            cursor.close()
            conn.close()
            flash("Exam Updated Successfully!", "success")
            return redirect(url_for('manage_exams'))
            
        except mysql.connector.Error as err:
            conn.rollback()
            cursor.close()
            conn.close()
            flash(f"Error: {err}", "danger")
            return redirect(url_for('edit_exam', exam_id=exam_id))
    
    # GET request - fetch exam data
    try:
        cursor.execute("SELECT exam_id, exam_title, subject_name, created_at, courses, time_limit FROM exam WHERE exam_id = %s", (exam_id,))
        exam = cursor.fetchone()
        
        if not exam:
            cursor.close()
            conn.close()
            flash("Exam not found!", "danger")
            return redirect(url_for('manage_exams'))
        
        cursor.execute("SELECT * FROM questions WHERE exam_id = %s ORDER BY question_id", (exam_id,))
        questions = cursor.fetchall()
        
        cursor.close()
        conn.close()
        return render_template('edit_exam.html', exam=exam, questions=questions)
        
    except mysql.connector.Error as err:
        cursor.close()
        conn.close()
        flash(f"Error: {err}", "danger")
        return redirect(url_for('manage_exams'))

# Toggle Score Visibility
@app.route('/admin/toggle_score_visibility/<int:exam_id>', methods=['POST'])
def toggle_score_visibility(exam_id):
    if 'admin_username' not in session:
        flash("Admin access required.", "danger")
        return redirect(url_for('home'))
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Get current visibility status
        cursor.execute("SELECT show_scores FROM exam WHERE exam_id = %s", (exam_id,))
        result = cursor.fetchone()
        
        if result:
            current_status = result[0] if result[0] is not None else 1
            new_status = 0 if current_status == 1 else 1
            
            # Update the visibility
            cursor.execute("UPDATE exam SET show_scores = %s WHERE exam_id = %s", (new_status, exam_id))
            conn.commit()
            
            status_text = "visible" if new_status == 1 else "hidden"
            flash(f"Score visibility updated! Scores are now {status_text} to students.", "success")
        else:
            flash("Exam not found!", "danger")
            
    except mysql.connector.Error as err:
        flash(f"Error updating visibility: {err}", "danger")
    finally:
        cursor.close()
        conn.close()
    
    return redirect(url_for('manage_exams'))

# Manage Registration Fields
@app.route('/admin/registration_fields')
def manage_registration_fields():
    if 'admin_username' not in session:
        flash("Admin access required.", "danger")
        return redirect(url_for('home'))
    
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM registration_fields ORDER BY field_order, field_id")
    fields = cursor.fetchall()
    cursor.close()
    conn.close()
    return render_template('manage_registration_fields.html', fields=fields)

@app.route('/admin/registration_fields/add', methods=['POST'])
def add_registration_field():
    if 'admin_username' not in session:
        return redirect(url_for('home'))
    
    field_name = request.form.get('field_name')
    field_label = request.form.get('field_label')
    field_type = request.form.get('field_type')
    field_options = request.form.get('field_options')
    is_required = request.form.get('is_required') == 'on'
    field_order = request.form.get('field_order', 0)
    
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("""
            INSERT INTO registration_fields (field_name, field_label, field_type, field_options, is_required, field_order, is_active)
            VALUES (%s, %s, %s, %s, %s, %s, TRUE)
        """, (field_name, field_label, field_type, field_options, is_required, field_order))
        conn.commit()
        flash("Field added successfully!", "success")
    except mysql.connector.Error as err:
        flash(f"Error: {err}", "danger")
    finally:
        cursor.close()
        conn.close()
    
    return redirect(url_for('manage_registration_fields'))

@app.route('/admin/registration_fields/delete/<int:field_id>', methods=['POST'])
def delete_registration_field(field_id):
    if 'admin_username' not in session:
        return redirect(url_for('home'))
    
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("DELETE FROM registration_fields WHERE field_id = %s", (field_id,))
        conn.commit()
        flash("Field deleted successfully!", "success")
    except mysql.connector.Error as err:
        flash(f"Error: {err}", "danger")
    finally:
        cursor.close()
        conn.close()
    
    return redirect(url_for('manage_registration_fields'))

@app.route('/admin/registration_fields/toggle/<int:field_id>', methods=['POST'])
def toggle_registration_field(field_id):
    if 'admin_username' not in session:
        return redirect(url_for('home'))
    
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("UPDATE registration_fields SET is_active = NOT is_active WHERE field_id = %s", (field_id,))
        conn.commit()
        flash("Field status updated!", "success")
    except mysql.connector.Error as err:
        flash(f"Error: {err}", "danger")
    finally:
        cursor.close()
        conn.close()
    
    return redirect(url_for('manage_registration_fields'))

# 📜 Conduct Exam (Admin)
@app.route('/admin/create_exam', methods=['GET', 'POST'])
def create_exam():
    if 'admin_username' not in session:
        flash("Admin access required.", "danger")
        return redirect(url_for('home'))
    
    # GET request - fetch available courses for the form
    if request.method == 'GET':
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT DISTINCT course FROM students WHERE course IS NOT NULL AND course != '' ORDER BY course")
        available_courses = [row[0] for row in cursor.fetchall()]
        
        # Add "All Courses" option at the beginning
        if available_courses:
            available_courses.insert(0, "All Courses")
        
        cursor.close()
        conn.close()
        return render_template('create_exam.html', available_courses=available_courses)
    
    if request.method == 'POST':
        exam_title = request.form.get("exam_title")
        subject = request.form.get("subject")
        time_limit = request.form.get("time_limit")  # Get time limit
        selected_courses = request.form.getlist("courses[]")  # Get selected courses
        start_datetime = request.form.get("start_datetime")  # Get start date/time
        end_datetime = request.form.get("end_datetime")  # Get end date/time
        
        # Handle question paper upload
        question_paper = request.files.get('question_paper')
        question_paper_path = None
        
        if question_paper and question_paper.filename and allowed_file(question_paper.filename):
            filename = secure_filename(question_paper.filename)
            unique_filename = f"exam_paper_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{filename}"
            question_paper_path = os.path.join('uploads', 'question_papers', unique_filename)
            os.makedirs(os.path.join('uploads', 'question_papers'), exist_ok=True)
            question_paper.save(question_paper_path)
            print(f"Question paper uploaded: {question_paper_path}")

        questions = request.form.getlist("question[]")
        question_types = request.form.getlist("question_type[]")
        options_a = request.form.getlist("option_a[]")
        options_b = request.form.getlist("option_b[]")
        options_c = request.form.getlist("option_c[]")
        options_d = request.form.getlist("option_d[]")
        correct_options = request.form.getlist("correct_option[]")
        tf_correct = request.form.getlist("tf_correct[]")
        explanations = request.form.getlist("explanation[]")
        question_images = request.files.getlist("question_image[]")
        
        # Debug logging
        print(f"DEBUG CREATE EXAM - Exam Title: {exam_title}, Subject: {subject}")
        print(f"DEBUG CREATE EXAM - Questions count: {len(questions)}")
        print(f"DEBUG CREATE EXAM - Questions: {questions}")
        print(f"DEBUG CREATE EXAM - Question Types: {question_types}")
        print(f"DEBUG CREATE EXAM - Form Data Keys: {list(request.form.keys())}")
        print(f"DEBUG CREATE EXAM - Files: {list(request.files.keys())}")
        
        # Check if no questions added
        if not questions or len(questions) == 0:
            flash("Error: You must add at least one question to the exam!", "danger")
            return redirect(url_for('create_exam'))

        conn = get_db_connection()
        cursor = conn.cursor()
        try:
            # Prepare courses string
            courses_str = ', '.join(selected_courses) if selected_courses else None
            
            # Convert time_limit to int or None
            time_limit_int = int(time_limit) if time_limit and time_limit.strip() else None
            
            # Convert datetime strings to proper format or None
            start_dt = start_datetime if start_datetime and start_datetime.strip() else None
            end_dt = end_datetime if end_datetime and end_datetime.strip() else None
            
            # Insert Exam Details with courses, time limit, question paper, and scheduling
            insert_exam_query = "INSERT INTO exam (exam_title, subject_name, courses, time_limit, question_paper_path, start_datetime, end_datetime) VALUES (%s, %s, %s, %s, %s, %s, %s)"
            cursor.execute(insert_exam_query, (exam_title, subject, courses_str, time_limit_int, question_paper_path, start_dt, end_dt))
            conn.commit()
            exam_id = cursor.lastrowid  # Get the newly inserted exam ID

            # Insert Questions into Database
            insert_question_query = """
            INSERT INTO questions (exam_id, question_text, question_type, option_a, option_b, option_c, option_d, correct_option, explanation, media_path) 
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """

            for i in range(len(questions)):
                q_type = question_types[i] if i < len(question_types) else 'mcq'
                media_path = None
                
                # Handle image upload for image_mcq questions
                if q_type == 'image_mcq' and i < len(question_images) and question_images[i].filename:
                    image_file = question_images[i]
                    filename = secure_filename(image_file.filename)
                    unique_filename = f"exam_{exam_id}_q_{i+1}_{filename}"
                    media_path = os.path.join('static', 'question_images', unique_filename)
                    os.makedirs(os.path.join('static', 'question_images'), exist_ok=True)
                    image_file.save(media_path)
                
                # Handle video upload for video_mcq questions
                elif q_type == 'video_mcq' and i < len(question_images) and question_images[i].filename:
                    video_file = question_images[i]
                    filename = secure_filename(video_file.filename)
                    unique_filename = f"exam_{exam_id}_q_{i+1}_video_{filename}"
                    media_path = os.path.join('static', 'question_images', unique_filename)
                    os.makedirs(os.path.join('static', 'question_images'), exist_ok=True)
                    video_file.save(media_path)
                
                # Handle different question types
                if q_type == 'video_response' or q_type == 'descriptive':
                    # No options for video response or descriptive
                    cursor.execute(insert_question_query, (
                        exam_id, questions[i], q_type, None, None, None, None, None, 
                        explanations[i] if i < len(explanations) else None, media_path
                    ))
                elif q_type == 'true_false':
                    # True/False questions
                    correct_ans = tf_correct[i] if i < len(tf_correct) else None
                    cursor.execute(insert_question_query, (
                        exam_id, questions[i], q_type, 'True', 'False', None, None, 
                        correct_ans, explanations[i] if i < len(explanations) else None, media_path
                    ))
                else:
                    # MCQ, Image MCQ, and Video MCQ questions
                    cursor.execute(insert_question_query, (
                        exam_id, questions[i], q_type, 
                        options_a[i] if i < len(options_a) else None,
                        options_b[i] if i < len(options_b) else None,
                        options_c[i] if i < len(options_c) else None,
                        options_d[i] if i < len(options_d) else None,
                        correct_options[i] if i < len(correct_options) else None,
                        explanations[i] if i < len(explanations) else None,
                        media_path
                    ))

            conn.commit()
            cursor.close()
            conn.close()
            
            # Create detailed success message
            courses_info = ', '.join(selected_courses) if selected_courses else 'all students'
            flash(f"🎯 Exam '{exam_title}' has been successfully assigned to {courses_info}!", "success")
            return redirect(url_for('admin_dashboard'))

        except mysql.connector.Error as err:
            conn.rollback()
            cursor.close()
            conn.close()
            flash(f"Error: {err}", "danger")

    return render_template('create_exam.html')


# � Download CSV Template for Questions
@app.route('/admin/download_template')
def download_template():
    """Download the sample questions template CSV file"""
    if 'admin_username' not in session:
        flash("Admin access required.", "danger")
        return redirect(url_for('home'))
    
    try:
        template_path = os.path.join(os.path.dirname(__file__), 'sample_questions_template.csv')
        return send_file(
            template_path,
            mimetype='text/csv',
            as_attachment=True,
            download_name='sample_questions_template.csv'
        )
    except Exception as e:
        flash(f"Error downloading template: {str(e)}", "danger")
        return redirect(url_for('create_exam'))

# �📤 Bulk Import Questions from CSV
@app.route('/admin/import_questions', methods=['POST'])
def import_questions():
    """Parse CSV file and return questions as JSON for bulk import"""
    if 'admin_username' not in session:
        return jsonify({'success': False, 'message': 'Admin access required'}), 403
    
    if 'csv_file' not in request.files:
        return jsonify({'success': False, 'message': 'No file uploaded'}), 400
    
    file = request.files['csv_file']
    
    if file.filename == '':
        return jsonify({'success': False, 'message': 'No file selected'}), 400
    
    if not file.filename.endswith('.csv'):
        return jsonify({'success': False, 'message': 'Please upload a CSV file'}), 400
    
    try:
        # Save file temporarily
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['QUESTION_TEMP_FOLDER'], filename)
        file.save(filepath)
        
        # Parse CSV file
        questions = []
        with open(filepath, 'r', encoding='utf-8') as csvfile:
            reader = csv.DictReader(csvfile)
            
            for row_num, row in enumerate(reader, start=2):  # Start at 2 (1 is header)
                try:
                    question_type = row.get('Question Type', '').strip().lower()
                    question_text = row.get('Question Text', '').strip()
                    
                    if not question_type or not question_text:
                        continue  # Skip empty rows
                    
                    question_data = {
                        'question_type': question_type,
                        'question_text': question_text,
                        'option_a': row.get('Option A', '').strip(),
                        'option_b': row.get('Option B', '').strip(),
                        'option_c': row.get('Option C', '').strip(),
                        'option_d': row.get('Option D', '').strip(),
                        'correct_answer': row.get('Correct Answer', '').strip(),
                        'explanation': row.get('Explanation/Instructions', '').strip()
                    }
                    
                    questions.append(question_data)
                    
                except Exception as e:
                    print(f"Error parsing row {row_num}: {str(e)}")
                    continue
        
        # Clean up temp file
        os.remove(filepath)
        
        if not questions:
            return jsonify({
                'success': False, 
                'message': 'No valid questions found in CSV. Please check the format.'
            }), 400
        
        return jsonify({
            'success': True,
            'message': f'Successfully imported {len(questions)} questions',
            'questions': questions
        }), 200
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error parsing CSV: {str(e)}'
        }), 500


# 📝 Attempt Exam (Student)
@app.route('/student/exam/<int:exam_id>')
def attempt_exam(exam_id):
    # ========== AUTHENTICATION & VALIDATION ==========
    if 'student_id' not in session:
        flash("Please log in first.", "warning")
        return redirect(url_for('unified_login'))
    
    student_id = session['student_id']
    
    # Validate exam_id is positive integer
    if not isinstance(exam_id, int) or exam_id <= 0:
        flash("Invalid exam ID!", "error")
        return redirect(url_for('student_dashboard'))
    
    conn = get_db_connection()
    if conn is None:
        flash("Database connection error. Please try again later.", "error")
        return redirect(url_for('student_dashboard'))
        
    cursor = conn.cursor()
    try:
        # ========== CHECK DUPLICATE ATTEMPT ==========
        cursor.execute("""
            SELECT performance_id, score, recorded_at 
            FROM student_performance 
            WHERE student_id = %s AND exam_id = %s
        """, (student_id, exam_id,))
        existing_attempt = cursor.fetchone()
        
        if existing_attempt:
            cursor.close()
            conn.close()
            flash("You have already taken this exam!", "warning")
            return redirect(url_for('student_dashboard'))
        
        # ========== FETCH & VALIDATE EXAM ==========
        cursor.execute("""
            SELECT exam_id, exam_title, subject_name, time_limit, 
                   start_datetime, end_datetime, question_paper_path
            FROM exam 
            WHERE exam_id = %s
        """, (exam_id,))
        exam = cursor.fetchone()
        
        if not exam:
            cursor.close()
            conn.close()
            flash("Exam not found!", "error")
            return redirect(url_for('student_dashboard'))
        
        # ========== VALIDATE EXAM SCHEDULE ==========
        if exam[4] and exam[5]:  # start_datetime and end_datetime exist
            from datetime import datetime
            current_time = datetime.now()
            start_datetime = exam[4]
            end_datetime = exam[5]
            
            # Check if exam hasn't started yet
            if current_time < start_datetime:
                cursor.close()
                conn.close()
                time_remaining = start_datetime - current_time
                flash(f"This exam is not available yet. It will start on {start_datetime.strftime('%B %d, %Y at %I:%M %p')}.", "warning")
                return redirect(url_for('student_dashboard'))
            
            # Check if exam has ended
            if current_time > end_datetime:
                cursor.close()
                conn.close()
                flash(f"This exam has ended. It closed on {end_datetime.strftime('%B %d, %Y at %I:%M %p')}.", "error")
                return redirect(url_for('student_dashboard'))
        
        # Validate time limit
        time_limit = exam[3]
        if time_limit is None or time_limit <= 0:
            cursor.close()
            conn.close()
            flash("Invalid exam time limit!", "error")
            return redirect(url_for('student_dashboard'))
        
        # ========== FETCH & VALIDATE QUESTIONS ==========
        cursor.execute("""
            SELECT question_id, exam_id, question_text, question_type,
                   option_a, option_b, option_c, option_d, correct_option,
                   explanation, media_path
            FROM questions 
            WHERE exam_id = %s
            ORDER BY question_id
        """, (exam_id,))
        questions = cursor.fetchall()
        
        print(f"[PRODUCTION] Exam ID: {exam_id} | Student: {student_id}")
        print(f"[PRODUCTION] Questions loaded: {len(questions) if questions else 0}")
        
        # Check if exam has questions
        if not questions or len(questions) == 0:
            cursor.close()
            conn.close()
            flash("This exam has no questions yet. Please contact your instructor.", "warning")
            return redirect(url_for('student_dashboard'))
        
        # ========== SHUFFLE QUESTIONS AND OPTIONS ==========
        questions_list = list(questions)
        random.shuffle(questions_list)  # Randomize question order
        
        # ========== SHUFFLE OPTIONS & MAP CORRECT ANSWERS ==========
        shuffled_questions = []
        option_mappings = {}  # question_id -> new_correct_option mapping
        
        for q in questions_list:
            q_list = list(q)
            question_id = q[0]
            question_type = q[3] if len(q) > 3 else 'mcq'
            
            # Shuffle options only for MCQ, Image MCQ, and True/False questions
            if question_type in ['mcq', 'image_mcq', 'true_false'] and len(q) > 5:
                try:
                    # Extract options safely
                    options = [
                        q[4] if len(q) > 4 else None, 
                        q[5] if len(q) > 5 else None, 
                        q[6] if len(q) > 6 else None, 
                        q[7] if len(q) > 7 else None
                    ]
                    correct_option = q[8] if len(q) > 8 and q[8] else 'A'  # Original correct option
                    
                    # Create option mapping with original letters
                    option_data = []
                    if options[0]: option_data.append(('A', options[0]))
                    if options[1]: option_data.append(('B', options[1]))
                    if options[2]: option_data.append(('C', options[2]))
                    if options[3]: option_data.append(('D', options[3]))
                    
                    if len(option_data) < 2:
                        # Not enough options to shuffle, keep original
                        shuffled_questions.append(q)
                        option_mappings[question_id] = correct_option
                        continue
                    
                    # Shuffle the options
                    random.shuffle(option_data)
                    
                    # Create mapping: new_position -> original_letter
                    new_letters = ['A', 'B', 'C', 'D']
                    for i, (original_letter, text) in enumerate(option_data):
                        if original_letter == correct_option:
                            # Store which new letter corresponds to correct answer
                            option_mappings[question_id] = new_letters[i]
                    
                    # Ensure q_list has enough elements
                    while len(q_list) < 12:
                        q_list.append(None)
                    
                    # Update question with shuffled options
                    q_list[4] = option_data[0][1] if len(option_data) > 0 else None
                    q_list[5] = option_data[1][1] if len(option_data) > 1 else None
                    q_list[6] = option_data[2][1] if len(option_data) > 2 else None
                    q_list[7] = option_data[3][1] if len(option_data) > 3 else None
                    
                    # Store the NEW correct option letter (after shuffle)
                    q_list[8] = option_mappings[question_id]
                    
                    shuffled_questions.append(tuple(q_list))
                except Exception as shuffle_err:
                    print(f"Error shuffling question {question_id}: {shuffle_err}")
                    # Keep original if shuffle fails
                    shuffled_questions.append(q)
                    option_mappings[question_id] = q[8] if len(q) > 8 else None
            else:
                # For non-MCQ questions, keep original
                shuffled_questions.append(q)
                option_mappings[question_id] = q[8] if len(q) > 8 else None
        
        # Store option mappings in session for answer validation
        session[f'exam_{exam_id}_mappings'] = option_mappings
        print(f"DEBUG SHUFFLE - Option mappings: {option_mappings}")
        
        return render_template('attempt_exam.html', exam=exam, questions=shuffled_questions)
        
    except mysql.connector.Error as db_err:
        cursor.close()
        conn.close()
        print(f"Database error in attempt_exam: {db_err}")
        flash("Database error occurred. Please try again.", "error")
        return redirect(url_for('student_dashboard'))
    except Exception as e:
        print(f"Error in attempt_exam: {e}")
        flash("An error occurred. Please try again.", "error")
        return redirect(url_for('student_dashboard'))

# ========== 📝 SUBMIT EXAM (PRODUCTION-LEVEL) ==========
@app.route('/student/exam/<int:exam_id>/submit', methods=['POST'])
def submit_exam(exam_id):
    # ========== AUTHENTICATION & SESSION VALIDATION ==========
    if 'student_id' not in session:
        flash("Session expired. Please log in again.", "warning")
        return redirect(url_for('unified_login'))

    student_id = session['student_id']
    student_name = session.get('student_name', 'Unknown')
    
    # Validate exam_id
    if not isinstance(exam_id, int) or exam_id <= 0:
        flash("Invalid exam submission!", "error")
        return redirect(url_for('student_dashboard'))

    print(f"[PRODUCTION-SUBMIT] Student: {student_id} | Exam: {exam_id}")
    print(f"[PRODUCTION-SUBMIT] Form keys: {list(request.form.keys())[:10]}...")  # Log first 10 keys

    # ========== DATABASE CONNECTION ==========
    conn = get_db_connection()
    if conn is None:
        print("[ERROR] Database connection failed (pool exhausted)")
        flash("Database connection error. Please try again.", "error")
        return redirect(url_for('student_dashboard'))
    
    cursor = conn.cursor()
    
    try:
        # ========== START TRANSACTION ==========
        conn.start_transaction()
        
        # ========== CHECK DUPLICATE SUBMISSION ==========
        cursor.execute("""
            SELECT performance_id, recorded_at 
            FROM student_performance 
            WHERE student_id = %s AND exam_id = %s
        """, (student_id, exam_id))
        
        existing_submission = cursor.fetchone()
        if existing_submission:
            conn.rollback()
            cursor.close()
            conn.close()
            flash("You have already submitted this exam!", "warning")
            return redirect(url_for('student_dashboard'))
        
        # ========== GET EXAM INFO & VALIDATE ==========
        cursor.execute("""
            SELECT exam_title, time_limit, end_datetime 
            FROM exam 
            WHERE exam_id = %s
        """, (exam_id,))
        exam = cursor.fetchone()
        
        if not exam:
            conn.rollback()
            cursor.close()
            conn.close()
            flash("Exam not found!", "error")
            return redirect(url_for('student_dashboard'))
        
        exam_title = exam[0]
        end_datetime = exam[2]
        
        # Check if exam is still open
        if end_datetime:
            from datetime import datetime
            if datetime.now() > end_datetime:
                conn.rollback()
                cursor.close()
                conn.close()
                flash("Exam time has expired!", "error")
                return redirect(url_for('student_dashboard'))

        # ========== GET ALL QUESTIONS WITH VALIDATION ==========
        cursor.execute("""
            SELECT question_id, question_type, correct_option, explanation
            FROM questions 
            WHERE exam_id = %s
            ORDER BY question_id
        """, (exam_id,))
        questions = cursor.fetchall()
        
        if not questions or len(questions) == 0:
            conn.rollback()
            cursor.close()
            conn.close()
            flash("No questions found for this exam!", "error")
            return redirect(url_for('student_dashboard'))
        
        # ========== RETRIEVE SHUFFLE MAPPINGS ==========
        option_mappings = session.get(f'exam_{exam_id}_mappings', {})
        print(f"[PRODUCTION-SUBMIT] Retrieved shuffle mappings for {len(option_mappings)} questions")
        
        # ========== PROCESS ANSWERS & CALCULATE SCORE ==========
        correct_count = 0
        total_mcq_questions = 0
        total_marks_obtained = 0
        total_possible_marks = 0
        
        # Process each question's answer
        for question in questions:
            question_id = question[0]
            question_type = question[1] if question[1] else 'mcq'
            correct_option = question[2]  # Original correct option from DB
            explanation = question[3] if len(question) > 3 else None
            marks = 1  # Default marks per question
            
            # Use shuffled correct answer if available
            if question_type in ['mcq', 'image_mcq', 'true_false']:
                # Check shuffle mapping (support both string and int keys)
                if str(question_id) in option_mappings:
                    correct_option = option_mappings[str(question_id)]
                elif question_id in option_mappings:
                    correct_option = option_mappings[question_id]
            
            # ========== HANDLE VIDEO RESPONSE ==========
            if question_type == 'video_response':
                video_submitted_key = f"video_submitted_{question_id}"
                video_submitted = request.form.get(video_submitted_key, "0")
                
                if video_submitted == "1":
                    print(f"[PRODUCTION-SUBMIT] Video response Q{question_id}: Submitted")
                else:
                    print(f"[PRODUCTION-SUBMIT] Video response Q{question_id}: NOT submitted")
            # ========== HANDLE DESCRIPTIVE ANSWER ==========
            elif question_type == 'descriptive':
                answer_key = f"answer_{question_id}"
                text_answer = request.form.get(answer_key, "").strip()
                
                # Validate answer length (prevent empty or too short answers)
                if text_answer and len(text_answer) >= 10:  # Minimum 10 characters
                    # Store descriptive response (no automatic scoring)
                    cursor.execute("""
                        INSERT INTO student_responses 
                        (student_id, exam_id, question_id, selected_option, is_correct, response_type, submitted_at)
                        VALUES (%s, %s, %s, %s, NULL, 'descriptive', NOW())
                    """, (student_id, exam_id, question_id, text_answer))
                    print(f"[PRODUCTION-SUBMIT] Descriptive answer Q{question_id}: {len(text_answer)} chars")
                elif not text_answer:
                    # Insert NULL answer for unanswered descriptive questions
                    cursor.execute("""
                        INSERT INTO student_responses 
                        (student_id, exam_id, question_id, selected_option, is_correct, response_type, submitted_at)
                        VALUES (%s, %s, %s, NULL, NULL, 'descriptive', NOW())
                    """, (student_id, exam_id, question_id))
                    print(f"[PRODUCTION-SUBMIT] Descriptive answer Q{question_id}: Unanswered")
            # ========== HANDLE MCQ/TRUE-FALSE/IMAGE MCQ ==========
            else:
                # Count this as an MCQ question
                total_mcq_questions += 1
                total_possible_marks += marks
                
                # Get student's answer for this question
                answer_key = f"answer_{question_id}"
                selected_answer = request.form.get(answer_key, "").strip()
                
                if selected_answer:
                    # Validate answer format
                    if question_type == 'true_false':
                        if selected_answer not in ['True', 'False']:
                            print(f"[WARNING] Invalid true/false answer Q{question_id}: {selected_answer}")
                            selected_answer = None
                    elif question_type in ['mcq', 'image_mcq']:
                        if selected_answer.upper() not in ['A', 'B', 'C', 'D']:
                            print(f"[WARNING] Invalid MCQ answer Q{question_id}: {selected_answer}")
                            selected_answer = None
                
                if selected_answer:
                    # Compare answers (case-insensitive for MCQ)
                    if question_type == 'true_false':
                        is_correct = 1 if selected_answer == correct_option else 0
                    else:
                        is_correct = 1 if selected_answer.upper() == correct_option.strip().upper() else 0
                    
                    correct_count += is_correct
                    if is_correct:
                        total_marks_obtained += marks
                    
                    # Store individual response
                    cursor.execute("""
                        INSERT INTO student_responses 
                        (student_id, exam_id, question_id, selected_option, is_correct, response_type, submitted_at)
                        VALUES (%s, %s, %s, %s, %s, %s, NOW())
                    """, (student_id, exam_id, question_id, selected_answer, is_correct, question_type))
                    
                    print(f"[PRODUCTION-SUBMIT] Q{question_id}: Answer={selected_answer}, Correct={correct_option}, Result={'✓' if is_correct else '✗'}")
                else:
                    # Store unanswered question
                    cursor.execute("""
                        INSERT INTO student_responses 
                        (student_id, exam_id, question_id, selected_option, is_correct, response_type, submitted_at)
                        VALUES (%s, %s, %s, NULL, 0, %s, NOW())
                    """, (student_id, exam_id, question_id, question_type))
                    print(f"[PRODUCTION-SUBMIT] Q{question_id}: Unanswered")
        # ========== CALCULATE FINAL SCORE ==========
        incorrect_count = total_mcq_questions - correct_count
        score = (correct_count / total_mcq_questions * 100) if total_mcq_questions > 0 else 0
        total_questions = len(questions)
        
        print(f"[PRODUCTION-SUBMIT] Total: {total_questions} | MCQ: {total_mcq_questions} | Correct: {correct_count} | Score: {score:.2f}%")
        
        # ========== UPDATE STUDENT PERFORMANCE ==========
        cursor.execute("""
            INSERT INTO student_performance 
            (student_name, student_id, exam_id, total_questions, correct_answers, incorrect_answers, score, recorded_at)
            VALUES (%s, %s, %s, %s, %s, %s, %s, NOW())
            ON DUPLICATE KEY UPDATE 
                total_questions = VALUES(total_questions), 
                correct_answers = VALUES(correct_answers),
                incorrect_answers = VALUES(incorrect_answers),
                score = VALUES(score),
                recorded_at = NOW()
        """, (student_name, student_id, exam_id, total_questions, correct_count, incorrect_count, score))

        # ========== COMMIT TRANSACTION ==========
        conn.commit()
        print(f"[PRODUCTION-SUBMIT] Transaction committed successfully")
        
        # ========== CLEANUP ==========
        cursor.close()
        conn.close()
        
        # Clear shuffle mappings from session
        session.pop(f'exam_{exam_id}_mappings', None)

        # ========== STORE RESULTS IN SESSION ==========
        session['exam_result'] = {
            "exam_title": exam_title,
            "total_questions": total_questions,
            "correct_answers": correct_count,
            "incorrect_answers": incorrect_count,
            "score": score
        }
        
        print(f"[PRODUCTION-SUBMIT] Submission complete! Redirecting to results...")
        return redirect(url_for('exam_result', exam_id=exam_id))
    
    # ========== ERROR HANDLING ==========
    except mysql.connector.Error as db_err:
        print(f"[ERROR-DB] submit_exam: {db_err}")
        print(f"[ERROR-DB] Error Code: {db_err.errno}")
        conn.rollback()
        cursor.close()
        conn.close()
        flash("Database error occurred while submitting exam. Please try again.", "error")
        return redirect(url_for('student_dashboard'))
    
    except ValueError as val_err:
        print(f"[ERROR-VALUE] submit_exam: {val_err}")
        conn.rollback()
        cursor.close()
        conn.close()
        flash("Invalid data submitted. Please check your answers.", "error")
        return redirect(url_for('attempt_exam', exam_id=exam_id))
    
    except Exception as e:
        print(f"[ERROR-GENERAL] submit_exam: {e}")
        import traceback
        traceback.print_exc()
        
        try:
            conn.rollback()
            cursor.close()
            conn.close()
        except:
            pass
        
        flash("An unexpected error occurred. Please try again.", "error")
        return redirect(url_for('student_dashboard'))# 🎥 Upload Video Response
@app.route('/upload_video_response', methods=['POST'])
def upload_video_response():
    if 'student_id' not in session:
        return jsonify({'success': False, 'message': 'Not logged in'}), 401
    
    try:
        # Get fresh database connection
        conn = get_db_connection()
        cursor = conn.cursor()
        
        student_id = session['student_id']
        exam_id = request.form.get('exam_id')
        question_id = request.form.get('question_id')
        attempt_number = request.form.get('attempt_number', '1')
        
        print(f"DEBUG: Upload video - student={student_id}, exam={exam_id}, question={question_id}, attempt={attempt_number}")
        
        if 'video' not in request.files:
            return jsonify({'success': False, 'message': 'No video file'}), 400
        
        video_file = request.files['video']
        
        if video_file.filename == '':
            return jsonify({'success': False, 'message': 'Empty filename'}), 400
        
        # Create filename: studentID_examID_questionID_attempt.webm
        filename = f"student_{student_id}_exam_{exam_id}_q_{question_id}_attempt_{attempt_number}.webm"
        filepath = os.path.join(app.config['STUDENT_RESPONSES_FOLDER'], filename)
        
        # Save video file
        video_file.save(filepath)
        print(f"DEBUG: Video saved to {filepath}")
        
        # Get video duration (if provided)
        duration = request.form.get('duration', 0)
        
        # Check if response already exists for this student/exam/question
        cursor.execute("""
            SELECT response_id FROM student_responses 
            WHERE student_id = %s AND exam_id = %s AND question_id = %s
        """, (student_id, exam_id, question_id))
        
        existing = cursor.fetchone()
        
        if existing:
            # Update existing record
            cursor.execute("""
                UPDATE student_responses 
                SET media_path = %s, duration = %s, response_type = 'video', 
                    submitted_at = NOW()
                WHERE student_id = %s AND exam_id = %s AND question_id = %s
            """, (filename, duration, student_id, exam_id, question_id))
            print(f"DEBUG: Updated existing response_id={existing[0]}")
        else:
            # Insert new record (selected_option is NULL for video responses)
            cursor.execute("""
                INSERT INTO student_responses 
                (student_id, exam_id, question_id, response_type, media_path, duration, is_correct)
                VALUES (%s, %s, %s, 'video', %s, %s, NULL)
            """, (student_id, exam_id, question_id, filename, duration))
            print(f"DEBUG: Inserted new response")
        
        conn.commit()
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True, 
            'message': 'Video uploaded successfully',
            'filename': filename
        }), 200
        
    except Exception as e:
        print(f"ERROR in upload_video_response: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


# � Download Question Paper (Student - After Submission)
@app.route('/download_question_paper/<int:exam_id>')
def download_question_paper(exam_id):
    """Allow students to download question paper after submitting exam"""
    if 'student_id' not in session:
        flash("Please log in to access this resource.", "warning")
        return redirect(url_for('home'))
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Check if student has submitted this exam
        student_id = session.get('student_id')
        cursor.execute("""
            SELECT COUNT(*) FROM student_performance 
            WHERE student_id = %s AND exam_id = %s
        """, (student_id, exam_id))
        has_submitted = cursor.fetchone()[0] > 0
        
        if not has_submitted:
            flash("You must submit the exam before downloading the question paper.", "warning")
            return redirect(url_for('student_dashboard'))
        
        # Get question paper path
        cursor.execute("SELECT question_paper_path, exam_title FROM exam WHERE exam_id = %s", (exam_id,))
        result = cursor.fetchone()
        
        if not result or not result[0]:
            flash("No question paper available for this exam.", "info")
            return redirect(url_for('student_dashboard'))
        
        question_paper_path = result[0]
        exam_title = result[1]
        
        # Convert relative path to absolute path
        absolute_path = os.path.abspath(question_paper_path)
        
        # Check if file exists
        if not os.path.exists(absolute_path):
            flash("Question paper file not found.", "danger")
            return redirect(url_for('student_dashboard'))
        
        # Send file for download
        return send_file(
            absolute_path,
            as_attachment=True,
            download_name=f"{exam_title}_Question_Paper.{question_paper_path.split('.')[-1]}"
        )
        
    except Exception as e:
        flash(f"Error downloading question paper: {str(e)}", "danger")
        return redirect(url_for('student_dashboard'))
    finally:
        cursor.close()
        conn.close()


# 📊 Student Exam Results
@app.route('/student/exam/<int:exam_id>/result')
def exam_result(exam_id):
    exam_result = session.get('exam_result', None)

    print(f"DEBUG: Session Exam Result: {exam_result}")  # Debugging

    if not exam_result:
        flash("No exam result found! Please attempt an exam first.", "error")
        return redirect(url_for('student_dashboard'))
    
    # Check if admin has enabled score visibility for this exam
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT show_scores FROM exam WHERE exam_id = %s", (exam_id,))
    result = cursor.fetchone()
    cursor.close()
    conn.close()
    
    show_scores = result[0] if result and result[0] is not None else 1  # Default to showing scores
    
    return render_template('exam_result.html', **exam_result, exam_id=exam_id, show_scores=show_scores)

@app.route('/admin/student_performance')
def student_performance():
    if 'admin_username' not in session:
        flash("Admin access required.", "danger")
        return redirect(url_for('home'))
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT sp.student_id, s.name, e.exam_title, sp.total_questions, 
               sp.correct_answers, sp.incorrect_answers, sp.score
        FROM student_performance sp
        JOIN students s ON sp.student_id = s.student_id
        JOIN exam e ON sp.exam_id = e.exam_id
    """)
    
    performance_records = cursor.fetchall()
    
    cursor.close()
    conn.close()
    return render_template('student_performance.html', records=performance_records)

# 📋 Admin - View Pending Student Registrations
@app.route('/admin/pending_students')
def pending_students():
    if 'admin_username' not in session:
        flash("Admin access required.", "danger")
        return redirect(url_for('home'))
    
    conn = get_db_connection()
    cursor = conn.cursor()
    # Get all students ordered by name to create consistent S.No mapping
    cursor.execute("SELECT student_id, name FROM students ORDER BY name")
    all_students_for_mapping = cursor.fetchall()
    student_id_to_sno = {student[0]: idx + 1 for idx, student in enumerate(all_students_for_mapping)}
    
    cursor.execute("SELECT student_id, name, email, course, status, created_at FROM students WHERE status='pending' ORDER BY created_at DESC")
    pending_students = cursor.fetchall()
    
    cursor.execute("SELECT student_id, name, email, course, status, created_at FROM students WHERE status IN ('approved', 'rejected') ORDER BY created_at DESC")
    all_students = cursor.fetchall()
    
    cursor.close()
    conn.close()
    return render_template('pending_students.html', 
                         pending_students=pending_students, 
                         all_students=all_students,
                         student_id_to_sno=student_id_to_sno)

# ✅ Admin - Approve Student
@app.route('/admin/approve_student/<int:student_id>', methods=['POST'])
def approve_student(student_id):
    if 'admin_username' not in session:
        flash("Admin access required.", "danger")
        return redirect(url_for('home'))
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("UPDATE students SET status='approved' WHERE student_id=%s", (student_id,))
    conn.commit()
    cursor.close()
    conn.close()
    flash("Student approved successfully!", "success")
    return redirect(url_for('pending_students'))

# ❌ Admin - Reject Student
@app.route('/admin/reject_student/<int:student_id>', methods=['POST'])
def reject_student(student_id):
    if 'admin_username' not in session:
        flash("Admin access required.", "danger")
        return redirect(url_for('home'))
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("UPDATE students SET status='rejected' WHERE student_id=%s", (student_id,))
    conn.commit()
    cursor.close()
    conn.close()
    flash("Student rejected!", "info")
    return redirect(url_for('pending_students'))

# � Admin - Reset Student Password
@app.route('/admin/reset_student_password/<int:student_id>', methods=['POST'])
def reset_student_password(student_id):
    if 'admin_username' not in session:
        return jsonify({'success': False, 'message': 'Admin access required'}), 403
    
    new_password = request.form.get('new_password')
    if not new_password or len(new_password) < 6:
        return jsonify({'success': False, 'message': 'Password must be at least 6 characters'}), 400
    
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        # Update student password
        hashed_password = generate_password_hash(new_password)
        cursor.execute("UPDATE students SET password=%s WHERE student_id=%s", (hashed_password, student_id))
        conn.commit()
        return jsonify({'success': True, 'new_password': new_password})
    except Exception as e:
        conn.rollback()
        print(f"Error resetting password: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        cursor.close()
        conn.close()

# �🗑️ Admin - Delete Student
@app.route('/admin/delete_student/<int:student_id>', methods=['POST'])
def delete_student(student_id):
    if 'admin_username' not in session:
        flash("Admin access required.", "danger")
        return redirect(url_for('home'))
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        # Delete related records first (foreign key constraints)
        cursor.execute("DELETE FROM student_responses WHERE student_id=%s", (student_id,))
        cursor.execute("DELETE FROM student_performance WHERE student_id=%s", (student_id,))
        cursor.execute("DELETE FROM proctor_logs WHERE student_id=%s", (student_id,))
        # Skip results table - it doesn't have student_id column
        cursor.execute("DELETE FROM coding_results WHERE student_id=%s", (student_id,))
        # Finally delete the student
        cursor.execute("DELETE FROM students WHERE student_id=%s", (student_id,))
        conn.commit()
        flash("Student deleted successfully!", "success")
    except Exception as e:
        conn.rollback()
        print(f"Error deleting student: {str(e)}")
        flash(f"Error deleting student: {str(e)}", "danger")
    finally:
        cursor.close()
        conn.close()
    return redirect(url_for('pending_students'))

# 👥 Admin - View All Students Information
@app.route('/admin/all_students_info')
def all_students_info():
    if 'admin_username' not in session:
        flash("Admin access required.", "danger")
        return redirect(url_for('home'))
    
    conn = get_db_connection()
    cursor = conn.cursor()
    # Get all students
    cursor.execute("SELECT student_id, name, email, course, status, created_at FROM students ORDER BY created_at DESC")
    students = cursor.fetchall()
    
    # Calculate statistics
    total_students = len(students)
    approved_count = sum(1 for s in students if s[4] == 'approved')
    pending_count = sum(1 for s in students if s[4] == 'pending')
    rejected_count = sum(1 for s in students if s[4] == 'rejected')
    
    # Get unique courses for filter
    cursor.execute("SELECT DISTINCT course FROM students ORDER BY course")
    courses = [row[0] for row in cursor.fetchall()]
    
    cursor.close()
    conn.close()
    return render_template('all_students_info.html',
                         students=students,
                         total_students=total_students,
                         approved_count=approved_count,
                         pending_count=pending_count,
                         rejected_count=rejected_count,
                         courses=courses)

# 📥 Bulk Import Students - Display Upload Page
@app.route('/admin/bulk_import_students')
def bulk_import_students():
    if 'admin_username' not in session:
        flash("Admin access required.", "danger")
        return redirect(url_for('home'))
    return render_template('bulk_import_students.html')

# 📥 Download Sample CSV Template
@app.route('/admin/download_sample_csv')
def download_sample_csv():
    if 'admin_username' not in session:
        flash("Admin access required.", "danger")
        return redirect(url_for('home'))
    
    # Create sample CSV content
    csv_content = "name,email,course,password\n"
    csv_content += "John Doe,john@test.com,Computer Science,john123\n"
    csv_content += "Jane Smith,jane@test.com,Mathematics,jane123\n"
    csv_content += "Bob Wilson,bob@test.com,Physics,bob123\n"
    
    # Create in-memory file
    output = io.BytesIO()
    output.write(csv_content.encode('utf-8'))
    output.seek(0)
    
    return send_file(
        output,
        mimetype='text/csv',
        as_attachment=True,
        download_name='student_import_sample.csv'
    )

# 📥 Process Bulk Import
@app.route('/admin/process_bulk_import', methods=['POST'])
def process_bulk_import():
    if 'admin_username' not in session:
        flash("Admin access required.", "danger")
        return redirect(url_for('home'))
    
    if 'csv_file' not in request.files:
        flash("No file uploaded!", "danger")
        return redirect(url_for('bulk_import_students'))
    
    file = request.files['csv_file']
    if file.filename == '':
        flash("No file selected!", "danger")
        return redirect(url_for('bulk_import_students'))
    
    if not file.filename.endswith('.csv'):
        flash("Please upload a CSV file!", "danger")
        return redirect(url_for('bulk_import_students'))
    
    # Get status preference (approved or pending)
    status = request.form.get('status', 'approved')
    
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        # Read CSV file
        csv_data = file.read().decode('utf-8').splitlines()
        csv_reader = csv.DictReader(csv_data)
        
        # Validate headers
        required_headers = ['name', 'email', 'course', 'password']
        if not all(header in csv_reader.fieldnames for header in required_headers):
            cursor.close()
            conn.close()
            flash(f"CSV must have columns: {', '.join(required_headers)}", "danger")
            return redirect(url_for('bulk_import_students'))
        
        # Store results for credential file
        import_results = []
        success_count = 0
        skip_count = 0
        error_count = 0
        
        for row in csv_reader:
            try:
                name = row['name'].strip()
                email = row['email'].strip().lower()
                course = row['course'].strip()
                password = row['password'].strip()
                
                # Validate required fields
                if not all([name, email, course, password]):
                    import_results.append({
                        'name': name or 'N/A',
                        'email': email or 'N/A',
                        'course': course or 'N/A',
                        'password': '***',
                        'status': 'Error',
                        'message': 'Missing required fields'
                    })
                    error_count += 1
                    continue
                
                # Check if email already exists
                cursor.execute("SELECT student_id FROM students WHERE email = %s", (email,))
                existing = cursor.fetchone()
                
                if existing:
                    import_results.append({
                        'name': name,
                        'email': email,
                        'course': course,
                        'password': '***',
                        'status': 'Skipped',
                        'message': 'Email already exists'
                    })
                    skip_count += 1
                    continue
                
                # Hash password and insert
                hashed_password = generate_password_hash(password)
                cursor.execute(
                    "INSERT INTO students (name, email, password, course, status) VALUES (%s, %s, %s, %s, %s)",
                    (name, email, hashed_password, course, status)
                )
                conn.commit()
                
                import_results.append({
                    'name': name,
                    'email': email,
                    'course': course,
                    'password': password,  # Store plain password for credential file
                    'status': 'Success',
                    'message': f'Account created ({status})'
                })
                success_count += 1
                
            except Exception as e:
                import_results.append({
                    'name': row.get('name', 'N/A'),
                    'email': row.get('email', 'N/A'),
                    'course': row.get('course', 'N/A'),
                    'password': '***',
                    'status': 'Error',
                    'message': str(e)
                })
                error_count += 1
                conn.rollback()
        
        # Store results in session for download
        session['import_results'] = import_results
        session['import_stats'] = {
            'success': success_count,
            'skipped': skip_count,
            'errors': error_count,
            'total': len(import_results)
        }
        
        cursor.close()
        conn.close()
        flash(f"Import Complete! ✅ Success: {success_count} | ⚠️ Skipped: {skip_count} | ❌ Errors: {error_count}", "success")
        return redirect(url_for('bulk_import_results'))
        
    except Exception as e:
        cursor.close()
        conn.close()
        flash(f"Error processing CSV: {str(e)}", "danger")
        return redirect(url_for('bulk_import_students'))

# 📊 Bulk Import Results & Download Credentials
@app.route('/admin/bulk_import_results')
def bulk_import_results():
    if 'admin_username' not in session:
        flash("Admin access required.", "danger")
        return redirect(url_for('home'))
    
    import_results = session.get('import_results', [])
    import_stats = session.get('import_stats', {})
    
    if not import_results:
        flash("No import results found!", "warning")
        return redirect(url_for('bulk_import_students'))
    
    return render_template('bulk_import_results.html', 
                         results=import_results, 
                         stats=import_stats)

# 📥 Download Credentials Excel File
@app.route('/admin/download_credentials_excel')
def download_credentials_excel():
    if 'admin_username' not in session:
        flash("Admin access required.", "danger")
        return redirect(url_for('home'))
    
    import_results = session.get('import_results', [])
    
    if not import_results:
        flash("No import results found!", "warning")
        return redirect(url_for('bulk_import_students'))
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Student Credentials"
    
    # Define styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    success_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    warning_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    error_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ['Name', 'Email', 'Course', 'Password', 'Status', 'Message']
    ws.append(headers)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Data rows
    for result in import_results:
        row = [
            result['name'],
            result['email'],
            result['course'],
            result['password'],
            result['status'],
            result['message']
        ]
        ws.append(row)
        
        row_num = ws.max_row
        
        # Apply color coding based on status
        status = result['status']
        if status == 'Success':
            fill = success_fill
        elif status == 'Skipped':
            fill = warning_fill
        else:
            fill = error_fill
        
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # Auto-adjust column widths
    for col_num, header in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        max_length = len(header)
        for row in ws.iter_rows(min_col=col_num, max_col=col_num):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = min(max_length + 3, 50)
    
    # Save to memory
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"Student_Credentials_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )
    return redirect(url_for('pending_students'))


# 🎥 AI PROCTORING - Enhanced Face Detection
face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')
eye_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_eye.xml')

# Temporal tracking for improved accuracy
proctor_state = {}
# Format: {student_id: {'no_face_count': 0, 'multiple_faces_count': 0, 'last_alert_time': 0, 'total_frames': 0}}

NO_FACE_THRESHOLD = 3  # Need 3 consecutive frames without face to alert
MULTIPLE_FACES_THRESHOLD = 2  # Need 2 consecutive frames with multiple faces
ALERT_COOLDOWN = 10  # Minimum 10 seconds between same type alerts

@socketio.on('proctor_frame')
def handle_proctor_frame(data):
    """Receive webcam frame and perform enhanced face detection with temporal filtering"""
    try:
        student_id = session.get('student_id')
        exam_id = data.get('exam_id')
        
        # Initialize student state if not exists
        if student_id not in proctor_state:
            proctor_state[student_id] = {
                'no_face_count': 0,
                'multiple_faces_count': 0,
                'last_no_face_alert': 0,
                'last_multiple_alert': 0,
                'total_frames': 0,
                'good_frames': 0
            }
        
        state = proctor_state[student_id]
        state['total_frames'] += 1
        
        # Decode base64 image
        img_data = data['frame'].split(',')[1]
        img_bytes = base64.b64decode(img_data)
        nparr = np.frombuffer(img_bytes, np.uint8)
        img = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
        
        if img is None:
            return
        
        # Convert to grayscale for face detection
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        
        # Apply histogram equalization for better detection in varying lighting
        gray = cv2.equalizeHist(gray)
        
        # Improved face detection with optimized parameters
        # scaleFactor=1.1 (more thorough), minNeighbors=4 (balanced accuracy)
        # minSize prevents tiny false detections
        faces = face_cascade.detectMultiScale(
            gray,
            scaleFactor=1.1,
            minNeighbors=4,
            minSize=(80, 80),  # Minimum face size
            flags=cv2.CASCADE_SCALE_IMAGE
        )
        
        face_count = len(faces)
        current_time = time.time()
        
        # Enhanced validation: check for eyes within detected faces for higher confidence
        verified_faces = 0
        for (x, y, w, h) in faces:
            roi_gray = gray[y:y+h, x:x+w]
            eyes = eye_cascade.detectMultiScale(roi_gray, scaleFactor=1.1, minNeighbors=3)
            # If eyes detected, it's very likely a real face
            if len(eyes) >= 1:  # At least one eye detected
                verified_faces += 1
            else:
                # Even without eye detection, if face is large enough, count it
                if w * h > 12000:  # Large face area threshold
                    verified_faces += 1
        
        # Use verified face count for more accurate detection
        face_count = verified_faces
        
        # TEMPORAL FILTERING: Track consecutive detections
        if face_count == 0:
            state['no_face_count'] += 1
            state['multiple_faces_count'] = 0  # Reset other counter
            
            # Only alert after consecutive frames without face
            if state['no_face_count'] >= NO_FACE_THRESHOLD:
                # Check cooldown period
                if current_time - state['last_no_face_alert'] >= ALERT_COOLDOWN:
                    log_proctor_event(student_id, exam_id, 'no_face', 
                                    f'No face detected for {state["no_face_count"]} consecutive frames')
                    socketio.emit('proctor_alert', {
                        'type': 'warning',
                        'message': '⚠️ Please ensure your face is visible in the camera!'
                    })
                    state['last_no_face_alert'] = current_time
                    
        elif face_count > 1:
            state['multiple_faces_count'] += 1
            state['no_face_count'] = 0  # Reset other counter
            
            # Only alert after consecutive frames with multiple faces
            if state['multiple_faces_count'] >= MULTIPLE_FACES_THRESHOLD:
                # Check cooldown period
                if current_time - state['last_multiple_alert'] >= ALERT_COOLDOWN:
                    log_proctor_event(student_id, exam_id, 'multiple_faces', 
                                    f'{face_count} faces detected for {state["multiple_faces_count"]} consecutive frames')
                    socketio.emit('proctor_alert', {
                        'type': 'danger',
                        'message': f'🚨 Multiple faces detected ({face_count})! Only you should be visible.'
                    })
                    state['last_multiple_alert'] = current_time
                    
        else:  # Exactly 1 face - good state
            # Reset all counters when face is properly detected
            state['no_face_count'] = 0
            state['multiple_faces_count'] = 0
            state['good_frames'] += 1
        
        # Send compliance feedback every 50 frames
        if state['total_frames'] % 50 == 0:
            compliance_rate = (state['good_frames'] / state['total_frames']) * 100
            if compliance_rate >= 90:
                socketio.emit('proctor_feedback', {
                    'type': 'success',
                    'message': f'✓ Good compliance: {compliance_rate:.0f}%'
                })
        
    except Exception as e:
        print(f"Proctoring error: {e}")
        import traceback
        traceback.print_exc()

@socketio.on('proctor_event')
def handle_proctor_event(data):
    """Log proctoring events like tab switches"""
    try:
        student_id = session.get('student_id')
        exam_id = data.get('exam_id')
        event_type = data.get('event_type')
        description = data.get('description', '')
        
        log_proctor_event(student_id, exam_id, event_type, description)
        
    except Exception as e:
        print(f"Event logging error: {e}")

def log_proctor_event(student_id, exam_id, event_type, description):
    """Save proctoring event to database"""
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute(
            "INSERT INTO proctor_logs (student_id, exam_id, event_type, event_description) VALUES (%s, %s, %s, %s)",
            (student_id, exam_id, event_type, description)
        )
        conn.commit()
    except Exception as e:
        print(f"Database logging error: {e}")
    finally:
        cursor.close()
        conn.close()


# 🎬 Upload Media Response (Video/Audio/Image)
@app.route('/upload_media_response', methods=['POST'])
def upload_media_response():
    """Handle student media responses (video/audio answers)"""
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        student_id = session.get('student_id')
        exam_id = request.form.get('exam_id')
        question_id = request.form.get('question_id')
        response_type = request.form.get('response_type')  # 'video', 'audio', 'image'
        
        if 'media_file' not in request.files:
            cursor.close()
            conn.close()
            return jsonify({'success': False, 'message': 'No file provided'})
        
        file = request.files['media_file']
        
        if file and allowed_file(file.filename):
            filename = secure_filename(f"{student_id}_{exam_id}_{question_id}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.{file.filename.rsplit('.', 1)[1].lower()}")
            filepath = os.path.join(app.config['STUDENT_RESPONSES_FOLDER'], filename)
            file.save(filepath)
            
            # Save to database
            cursor.execute(
                "INSERT INTO student_responses (student_id, exam_id, question_id, response_type, media_path) VALUES (%s, %s, %s, %s, %s)",
                (student_id, exam_id, question_id, response_type, filepath)
            )
            conn.commit()
            cursor.close()
            conn.close()
            
            return jsonify({'success': True, 'message': 'Media uploaded successfully'})
        else:
            cursor.close()
            conn.close()
            return jsonify({'success': False, 'message': 'Invalid file type'})
            
    except Exception as e:
        cursor.close()
        conn.close()
        return jsonify({'success': False, 'message': str(e)})


# 📊 Student Report - Comprehensive Performance Report
@app.route('/student/report/<int:student_id>')
def student_report(student_id):
    """Generate comprehensive student report with exam history and proctoring logs"""
    if 'student_id' not in session and 'admin_username' not in session:
        flash("Access denied.", "danger")
        return redirect(url_for('home'))
    
    conn = get_db_connection()
    cursor = conn.cursor()
    # Get student info
    cursor.execute("SELECT student_id, name, email, course FROM students WHERE student_id = %s", (student_id,))
    student = cursor.fetchone()
    
    # Get exam history
    cursor.execute("""
        SELECT e.exam_title, sp.correct_answers, sp.total_questions, sp.score, sp.recorded_at
        FROM student_performance sp
        JOIN exam e ON sp.exam_id = e.exam_id
        WHERE sp.student_id = %s
        ORDER BY sp.recorded_at DESC
    """, (student_id,))
    exam_history = cursor.fetchall()
    
    # Get proctoring logs
    cursor.execute("""
        SELECT pl.event_type, pl.event_description, pl.timestamp, e.exam_title
        FROM proctor_logs pl
        JOIN exam e ON pl.exam_id = e.exam_id
        WHERE pl.student_id = %s
        ORDER BY pl.timestamp DESC
    """, (student_id,))
    proctor_logs = cursor.fetchall()
    
    cursor.close()
    conn.close()
    return render_template('student_report.html', student=student, student_id=student_id, exam_history=exam_history, proctor_logs=proctor_logs)


# 📊 Export Single Student Report to Excel
@app.route('/student/report/<int:student_id>/export_excel')
def export_student_excel(student_id):
    """Export individual student report to Excel"""
    if 'student_id' not in session and 'admin_username' not in session:
        flash("Access denied.", "danger")
        return redirect(url_for('home'))
    
    # Get database connection
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Get student info
    cursor.execute("SELECT student_id, name, email, course FROM students WHERE student_id = %s", (student_id,))
    student = cursor.fetchone()
    
    if not student:
        flash("Student not found.", "error")
        return redirect(url_for('admin_dashboard'))
    
    # Get exam history
    cursor.execute("""
        SELECT e.exam_title, sp.correct_answers, sp.total_questions, sp.score, sp.recorded_at
        FROM student_performance sp
        JOIN exam e ON sp.exam_id = e.exam_id
        WHERE sp.student_id = %s
        ORDER BY sp.recorded_at DESC
    """, (student_id,))
    exam_history = cursor.fetchall()
    
    # Get proctoring logs
    cursor.execute("""
        SELECT pl.event_type, pl.event_description, pl.timestamp, e.exam_title
        FROM proctor_logs pl
        JOIN exam e ON pl.exam_id = e.exam_id
        WHERE pl.student_id = %s
        ORDER BY pl.timestamp DESC
    """, (student_id,))
    proctor_logs = cursor.fetchall()
    
    # Create Excel workbook
    wb = Workbook()
    
    # Define styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    title_font = Font(bold=True, size=14)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Sheet 1: Student Info & Summary
    ws1 = wb.active
    ws1.title = "Student Summary"
    
    # Title
    ws1['A1'] = "STUDENT PERFORMANCE REPORT"
    ws1['A1'].font = Font(bold=True, size=16)
    ws1.merge_cells('A1:D1')
    
    # Student Info
    ws1['A3'] = "Student ID:"
    ws1['B3'] = student[0]
    ws1['A4'] = "Name:"
    ws1['B4'] = student[1]
    ws1['A5'] = "Email:"
    ws1['B5'] = student[2]
    ws1['A6'] = "Course:"
    ws1['B6'] = student[3]
    
    for row in range(3, 7):
        ws1[f'A{row}'].font = Font(bold=True)
    
    # Statistics
    ws1['A8'] = "STATISTICS"
    ws1['A8'].font = title_font
    ws1['A9'] = "Total Exams Taken:"
    ws1['B9'] = len(exam_history)
    avg_score = sum([e[3] for e in exam_history]) / len(exam_history) if exam_history else 0
    ws1['A10'] = "Average Score:"
    ws1['B10'] = f"{avg_score:.2f}%"
    ws1['A11'] = "Total Proctoring Events:"
    ws1['B11'] = len(proctor_logs)
    
    for row in range(9, 12):
        ws1[f'A{row}'].font = Font(bold=True)
    
    # Sheet 2: Exam History
    ws2 = wb.create_sheet("Exam History")
    ws2['A1'] = "EXAM HISTORY"
    ws2['A1'].font = title_font
    ws2.merge_cells('A1:E1')
    
    headers = ["Exam Title", "Correct Answers", "Total Questions", "Score (%)", "Date"]
    for col, header in enumerate(headers, 1):
        cell = ws2.cell(row=3, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    for row_idx, exam in enumerate(exam_history, 4):
        ws2.cell(row=row_idx, column=1, value=exam[0]).border = border
        ws2.cell(row=row_idx, column=2, value=exam[1]).border = border
        ws2.cell(row=row_idx, column=3, value=exam[2]).border = border
        score_cell = ws2.cell(row=row_idx, column=4, value=f"{exam[3]:.2f}%")
        score_cell.border = border
        
        # Color code scores
        if exam[3] >= 80:
            score_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif exam[3] >= 50:
            score_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        else:
            score_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        ws2.cell(row=row_idx, column=5, value=str(exam[4])).border = border
    
    # Adjust column widths
    for col in range(1, 6):
        ws2.column_dimensions[get_column_letter(col)].width = 20
    
    # Sheet 3: Video Responses
    ws3 = wb.create_sheet("Video Responses")
    ws3['A1'] = "VIDEO RESPONSE RECORDINGS"
    ws3['A1'].font = title_font
    ws3.merge_cells('A1:F1')
    
    # Get video responses for this student
    try:
        cursor.execute("""
            SELECT q.question_text, e.exam_title, sr.media_path, sr.duration, sr.submitted_at, e.exam_id, q.question_id
            FROM student_responses sr
            JOIN questions q ON sr.question_id = q.question_id
            JOIN exam e ON sr.exam_id = e.exam_id
            WHERE sr.student_id = %s AND sr.response_type = 'video' AND sr.media_path IS NOT NULL
            ORDER BY sr.submitted_at DESC
        """, (student_id,))
        video_responses = cursor.fetchall()
        print(f"DEBUG EXPORT: Found {len(video_responses)} video responses for student {student_id}")
        for v in video_responses:
            print(f"  - Exam: {v[1]}, Question: {v[0][:30]}, File: {v[2]}")
    except Exception as e:
        print(f"ERROR fetching video responses: {e}")
        video_responses = []
    
    if video_responses:
        headers = ["Exam Title", "Question", "Video File", "Duration (sec)", "Submitted At", "File Path"]
        for col, header in enumerate(headers, 1):
            cell = ws3.cell(row=3, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        for row_idx, video in enumerate(video_responses, 4):
            ws3.cell(row=row_idx, column=1, value=video[1]).border = border  # Exam title
            ws3.cell(row=row_idx, column=2, value=video[0]).border = border  # Question text
            
            # Video filename with hyperlink
            video_filename = video[2]
            video_path = f"uploads/student_responses/{video_filename}"
            video_cell = ws3.cell(row=row_idx, column=3, value=video_filename)
            video_cell.border = border
            video_cell.font = Font(color="0000FF", underline="single")
            video_cell.hyperlink = f"file:///{os.path.abspath(video_path)}"
            
            ws3.cell(row=row_idx, column=4, value=video[3] if video[3] else 0).border = border  # Duration
            ws3.cell(row=row_idx, column=5, value=str(video[4])).border = border  # Submitted at
            ws3.cell(row=row_idx, column=6, value=video_path).border = border  # Full path
        
        # Adjust column widths
        ws3.column_dimensions['A'].width = 25
        ws3.column_dimensions['B'].width = 50
        ws3.column_dimensions['C'].width = 40
        ws3.column_dimensions['D'].width = 15
        ws3.column_dimensions['E'].width = 20
        ws3.column_dimensions['F'].width = 60
    else:
        ws3['A3'] = "No video responses recorded for this student."
        ws3['A3'].font = Font(italic=True, color="666666")
    
    # Sheet 4: Proctoring Logs
    ws4 = wb.create_sheet("Proctoring Logs")
    ws4['A1'] = "PROCTORING EVENTS LOG"
    ws4['A1'].font = title_font
    ws4.merge_cells('A1:D1')
    
    headers = ["Event Type", "Description", "Timestamp", "Exam"]
    for col, header in enumerate(headers, 1):
        cell = ws4.cell(row=3, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    for row_idx, log in enumerate(proctor_logs, 4):
        event_cell = ws4.cell(row=row_idx, column=1, value=log[0])
        event_cell.border = border
        
        # Color code event types
        if log[0] in ['no_face', 'multiple_faces']:
            event_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        elif log[0] == 'tab_switch':
            event_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        
        ws4.cell(row=row_idx, column=2, value=log[1]).border = border
        ws4.cell(row=row_idx, column=3, value=str(log[2])).border = border
        ws4.cell(row=row_idx, column=4, value=log[3]).border = border
    
    # Adjust column widths
    ws4.column_dimensions['A'].width = 20
    ws4.column_dimensions['B'].width = 40
    ws4.column_dimensions['C'].width = 20
    ws4.column_dimensions['D'].width = 25
    
    # Close database connection
    cursor.close()
    conn.close()
    
    # Save to BytesIO
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    filename = f"Student_Report_{student[1].replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    return send_file(
        excel_file,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


# 📊 Export All Students Report to Excel
@app.route('/admin/export_all_students_excel')
def export_all_students_excel():
    """Export complete class performance report to Excel"""
    if 'admin_username' not in session:
        flash("Please login as admin to export reports.", "danger")
        return redirect(url_for('home'))
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Get all students ordered by name
    cursor.execute("SELECT student_id, name, email, course, status FROM students ORDER BY name")
    students = cursor.fetchall()
    
    # Create a mapping of student_id to sequential number (S.No)
    student_id_to_sno = {student[0]: idx + 1 for idx, student in enumerate(students)}
    
    # Create Excel workbook
    wb = Workbook()
    
    # Define styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    title_font = Font(bold=True, size=16)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Sheet 1: All Students Summary
    ws1 = wb.active
    ws1.title = "Students Summary"
    
    ws1['A1'] = "ALL STUDENTS PERFORMANCE REPORT"
    ws1['A1'].font = title_font
    ws1.merge_cells('A1:G1')
    
    ws1['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws1.merge_cells('A2:G2')
    
    headers = ["S.No", "Name", "Email", "Course", "Status", "Exams Taken", "Average Score"]
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=4, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    for row_idx, student in enumerate(students, 5):
        ws1.cell(row=row_idx, column=1, value=row_idx - 4).border = border  # Sequential number
        ws1.cell(row=row_idx, column=2, value=student[1]).border = border
        ws1.cell(row=row_idx, column=3, value=student[2]).border = border
        ws1.cell(row=row_idx, column=4, value=student[3]).border = border
        
        status_cell = ws1.cell(row=row_idx, column=5, value=student[4])
        status_cell.border = border
        if student[4] == 'approved':
            status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif student[4] == 'pending':
            status_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        else:
            status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        # Get student performance
        cursor.execute("""
            SELECT COUNT(*), AVG(score)
            FROM student_performance
            WHERE student_id = %s
        """, (student[0],))
        perf = cursor.fetchone()
        
        ws1.cell(row=row_idx, column=6, value=perf[0] if perf[0] else 0).border = border
        avg_score = perf[1] if perf[1] else 0
        score_cell = ws1.cell(row=row_idx, column=7, value=f"{avg_score:.2f}%")
        score_cell.border = border
        
        if avg_score >= 80:
            score_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif avg_score >= 50:
            score_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        elif avg_score > 0:
            score_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    # Adjust column widths
    ws1.column_dimensions['A'].width = 8
    ws1.column_dimensions['B'].width = 25
    ws1.column_dimensions['C'].width = 30
    ws1.column_dimensions['D'].width = 20
    ws1.column_dimensions['E'].width = 12
    ws1.column_dimensions['F'].width = 15
    ws1.column_dimensions['G'].width = 15
    
    # Sheet 2: Detailed Performance
    ws2 = wb.create_sheet("Detailed Performance")
    ws2['A1'] = "DETAILED EXAM PERFORMANCE"
    ws2['A1'].font = title_font
    ws2.merge_cells('A1:G1')
    
    headers = ["S.No", "Student Name", "Exam Title", "Correct", "Total", "Score (%)", "Date"]
    for col, header in enumerate(headers, 1):
        cell = ws2.cell(row=3, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    # Get all performance records
    cursor.execute("""
        SELECT s.student_id, s.name, e.exam_title, sp.correct_answers, 
               sp.total_questions, sp.score, sp.recorded_at
        FROM student_performance sp
        JOIN students s ON sp.student_id = s.student_id
        JOIN exam e ON sp.exam_id = e.exam_id
        ORDER BY s.name, sp.recorded_at DESC
    """)
    all_performance = cursor.fetchall()
    
    for row_idx, record in enumerate(all_performance, 4):
        student_sno = student_id_to_sno.get(record[0], row_idx - 3)  # Use mapped S.No
        ws2.cell(row=row_idx, column=1, value=student_sno).border = border
        ws2.cell(row=row_idx, column=2, value=record[1]).border = border
        ws2.cell(row=row_idx, column=3, value=record[2]).border = border
        ws2.cell(row=row_idx, column=4, value=record[3]).border = border
        ws2.cell(row=row_idx, column=5, value=record[4]).border = border
        
        score_cell = ws2.cell(row=row_idx, column=6, value=f"{record[5]:.2f}%")
        score_cell.border = border
        
        if record[5] >= 80:
            score_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif record[5] >= 50:
            score_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        else:
            score_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        ws2.cell(row=row_idx, column=7, value=str(record[6])).border = border
    
    # Adjust column widths
    ws2.column_dimensions['A'].width = 12
    ws2.column_dimensions['B'].width = 25
    ws2.column_dimensions['C'].width = 25
    ws2.column_dimensions['D'].width = 12
    ws2.column_dimensions['E'].width = 12
    ws2.column_dimensions['F'].width = 12
    ws2.column_dimensions['G'].width = 20
    
    # Sheet 3: Proctoring Summary
    ws3 = wb.create_sheet("Proctoring Events")
    ws3['A1'] = "ALL PROCTORING EVENTS"
    ws3['A1'].font = title_font
    ws3.merge_cells('A1:E1')
    
    headers = ["S.No", "Student Name", "Exam", "Event Type", "Timestamp"]
    for col, header in enumerate(headers, 1):
        cell = ws3.cell(row=3, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    # Get all proctoring logs
    cursor.execute("""
        SELECT s.student_id, s.name, e.exam_title, pl.event_type, pl.timestamp
        FROM proctor_logs pl
        JOIN students s ON pl.student_id = s.student_id
        JOIN exam e ON pl.exam_id = e.exam_id
        ORDER BY pl.timestamp DESC
    """)
    all_logs = cursor.fetchall()
    
    for row_idx, log in enumerate(all_logs, 4):
        student_sno = student_id_to_sno.get(log[0], row_idx - 3)  # Use mapped S.No
        ws3.cell(row=row_idx, column=1, value=student_sno).border = border
        ws3.cell(row=row_idx, column=2, value=log[1]).border = border
        ws3.cell(row=row_idx, column=3, value=log[2]).border = border
        
        event_cell = ws3.cell(row=row_idx, column=4, value=log[3])
        event_cell.border = border
        
        if log[3] in ['no_face', 'multiple_faces']:
            event_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        elif log[3] == 'tab_switch':
            event_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        
        ws3.cell(row=row_idx, column=5, value=str(log[4])).border = border
    
    # Adjust column widths
    ws3.column_dimensions['A'].width = 12
    ws3.column_dimensions['B'].width = 25
    ws3.column_dimensions['C'].width = 25
    ws3.column_dimensions['D'].width = 20
    ws3.column_dimensions['E'].width = 20
    
    # Sheet 4: Video Responses
    ws4 = wb.create_sheet("Video Responses")
    ws4['A1'] = "ALL VIDEO RESPONSES"
    ws4['A1'].font = title_font
    ws4.merge_cells('A1:H1')
    
    headers = ["S.No", "Student Name", "Exam", "Question ID", "Video Path", "Duration (s)", "Submitted At"]
    for col, header in enumerate(headers, 1):
        cell = ws4.cell(row=3, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    # Get all video responses
    cursor.execute("""
        SELECT sr.response_id, sr.student_id, s.name, e.exam_title, 
               sr.question_id, sr.media_path, sr.duration, sr.submitted_at
        FROM student_responses sr
        JOIN students s ON sr.student_id = s.student_id
        JOIN exam e ON sr.exam_id = e.exam_id
        WHERE sr.response_type = 'video' AND sr.media_path IS NOT NULL
        ORDER BY sr.submitted_at DESC
    """)
    all_videos = cursor.fetchall()
    
    for row_idx, video in enumerate(all_videos, 4):
        student_sno = student_id_to_sno.get(video[1], row_idx - 3)  # Use mapped S.No from student_id
        ws4.cell(row=row_idx, column=1, value=student_sno).border = border
        ws4.cell(row=row_idx, column=2, value=video[2]).border = border
        ws4.cell(row=row_idx, column=3, value=video[3]).border = border
        ws4.cell(row=row_idx, column=4, value=video[4]).border = border
        
        # Video path with hyperlink
        video_path = video[5] if video[5] else "N/A"
        video_cell = ws4.cell(row=row_idx, column=5, value=video_path)
        video_cell.border = border
        video_cell.font = Font(color="0563C1", underline="single")
        
        # Add hyperlink if video exists
        if video[5] and os.path.exists(os.path.join('uploads', 'student_responses', os.path.basename(video[5]))):
            full_path = os.path.abspath(os.path.join('uploads', 'student_responses', os.path.basename(video[5])))
            video_cell.hyperlink = f"file:///{full_path}"
            video_cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        # Duration
        duration = video[6] if video[6] else 0
        ws4.cell(row=row_idx, column=6, value=duration).border = border
        
        # Timestamp
        ws4.cell(row=row_idx, column=7, value=str(video[7]) if video[7] else "N/A").border = border
    
    # Adjust column widths
    ws4.column_dimensions['A'].width = 8
    ws4.column_dimensions['B'].width = 25
    ws4.column_dimensions['C'].width = 25
    ws4.column_dimensions['D'].width = 12
    ws4.column_dimensions['E'].width = 50
    ws4.column_dimensions['F'].width = 15
    ws4.column_dimensions['G'].width = 20
    
    # Close database connection
    cursor.close()
    conn.close()
    
    # Save to BytesIO
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    filename = f"All_Students_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return send_file(
        excel_file,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


if __name__ == '__main__':
    socketio.run(app, host='0.0.0.0', port=5000, debug=True)
